# ============================================================
# AUTORÍA DEL CÓDIGO
# Autor: Aitor Santamaria Zuluaga
# Grado en Ingeniería en Organización Industrial
# Universidad Internacional de La Rioja (UNIR)
# JULIO DEL 2026 · Trabajo Fin de Grado (TFG)
# ============================================================

import os
import re
import random
import unicodedata
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

ARCHIVO_EXCEL_DATOS = "datos_centro.xlsx"
ARCHIVO_ASIGNACION = "asignacion_completa.xlsx"
ARCHIVO_HORARIOS = "archivoHorarios.xlsx"

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

def normalizar(texto):
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto


def misma_asignatura_grupo(a, b):
    return (
        normalizar(a["Asignatura"]) == normalizar(b["Asignatura"])
        and normalizar(a["Grupo"]) == normalizar(b["Grupo"])
    )


def obtener_grupos_asignaciones(asignaciones):
    return sorted(
        {a["Grupo"] for a in asignaciones},
        key=clave_orden_clase
    )

# ------------------------------------------------------------
# HORARIOS BASE
# ------------------------------------------------------------

def cargar_franjas_horarias():
    return {
        "ESO": [
            {"Sesion": 1, "Inicio": "09:00", "Fin": "09:55"},
            {"Sesion": 2, "Inicio": "09:55", "Fin": "10:50"},
            {"Sesion": 3, "Inicio": "11:20", "Fin": "12:15"},
            {"Sesion": 4, "Inicio": "12:15", "Fin": "13:10"},
            {"Sesion": 5, "Inicio": "14:30", "Fin": "15:30"},
            {"Sesion": 6, "Inicio": "15:30", "Fin": "16:30"},
        ],
        "BACH": [
            {"Sesion": 1, "Inicio": "08:05", "Fin": "09:00"},
            {"Sesion": 2, "Inicio": "09:00", "Fin": "09:55"},
            {"Sesion": 3, "Inicio": "09:55", "Fin": "10:50"},
            {"Sesion": 4, "Inicio": "11:20", "Fin": "12:15"},
            {"Sesion": 5, "Inicio": "12:15", "Fin": "13:10"},
            {"Sesion": 6, "Inicio": "13:10", "Fin": "14:05"},
            {"Sesion": 7, "Inicio": "14:05", "Fin": "15:00"},
        ]
    }


def obtener_etapa_grupo(grupo):
    grupo = str(grupo).strip().upper()

    if grupo.startswith("ESO"):
        return "ESO"

    if grupo.startswith("BACH"):
        return "BACH"

    return None


def obtener_sesiones_grupo_dia(grupo, dia, horarios):
    etapa = obtener_etapa_grupo(grupo)

    if etapa is None:
        return []

    sesiones = horarios[etapa]

    grupo_upper = str(grupo).strip().upper()
    dia_lower = str(dia).strip().lower()

    # 2º Bachillerato: jueves y viernes termina a las 14:05.
    if grupo_upper.startswith("BACH 2") and dia_lower in ["jueves", "viernes"]:
        return [s for s in sesiones if s["Sesion"] <= 6]

    # 1º Bachillerato: viernes termina a las 14:05.
    if grupo_upper.startswith("BACH 1") and dia_lower == "viernes":
        return [s for s in sesiones if s["Sesion"] <= 6]

    return sesiones


def obtener_franja_por_sesion(grupo, dia, sesion, horarios):
    sesiones = obtener_sesiones_grupo_dia(grupo, dia, horarios)

    for franja in sesiones:
        if franja["Sesion"] == sesion:
            return franja

    return None


def clave_horaria(grupo, dia, sesion, horarios):
    franja = obtener_franja_por_sesion(grupo, dia, sesion, horarios)

    if franja is None:
        return None

    return f"{franja['Inicio']}-{franja['Fin']}"


def hora_sesion(sesion, etapa, horarios):
    if etapa not in horarios:
        return ""

    for franja in horarios[etapa]:
        if franja["Sesion"] == sesion:
            return f"{franja['Inicio']}-{franja['Fin']}"

    return ""


# ------------------------------------------------------------
# LECTURA DE DATOS
# ------------------------------------------------------------

def cargar_aulas_excel():
    aulas = []

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return aulas

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "Aulas" not in wb.sheetnames:
        print("Error: No existe la hoja 'Aulas' en el archivo Excel.\n")
        return aulas

    ws = wb["Aulas"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    try:
        idx_asignatura = encabezados.index("asignatura")
        idx_grupo = encabezados.index("grupo")
        idx_tipo_aula = encabezados.index("tipoaula")
        idx_aula_fija = encabezados.index("aulafija")
        idx_aulas_posibles = encabezados.index("aulasposibles")
    except ValueError:
        print("Error: La hoja 'Aulas' debe tener las columnas:")
        print("'Asignatura', 'Grupo', 'TipoAula', 'AulaFija', 'AulasPosibles'.\n")
        return aulas

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        asignatura = str(row[idx_asignatura]).strip()
        grupo = str(row[idx_grupo]).strip()
        tipo_aula = str(row[idx_tipo_aula]).strip()

        aula_fija = ""
        if row[idx_aula_fija] is not None:
            aula_fija = str(row[idx_aula_fija]).strip()

        aulas_posibles = []
        if row[idx_aulas_posibles] is not None:
            aulas_posibles = [
                a.strip()
                for a in re.split(r"[;,]", str(row[idx_aulas_posibles]))
                if a.strip() != ""
            ]

        aulas.append({
            "Asignatura": asignatura,
            "Grupo": grupo,
            "TipoAula": tipo_aula,
            "AulaFija": aula_fija,
            "AulasPosibles": aulas_posibles
        })

    print(f"{len(aulas)} configuraciones de aula cargadas correctamente desde la hoja 'Aulas'.\n")
    return aulas


def cargar_desdobles_excel():
    desdobles = {}

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return desdobles

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "Desdobles" not in wb.sheetnames:
        print("No existe la hoja 'Desdobles' en el archivo Excel.\n")
        return desdobles

    ws = wb["Desdobles"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    try:
        idx_bloque = encabezados.index("bloque")
        idx_tipo = encabezados.index("tipo")
        idx_asignatura = encabezados.index("asignatura")
        idx_grupo = encabezados.index("grupo")
    except ValueError:
        print("Error: La hoja 'Desdobles' debe tener las columnas:")
        print("'Bloque', 'Tipo', 'Asignatura', 'Grupo'.\n")
        return desdobles

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        bloque = str(row[idx_bloque]).strip()
        tipo = str(row[idx_tipo]).strip().upper()
        asignatura = str(row[idx_asignatura]).strip()
        grupo = str(row[idx_grupo]).strip()

        if tipo not in ["C", "G"]:
            tipo = "G"

        if bloque not in desdobles:
            desdobles[bloque] = {
                "Tipo": tipo,
                "Materias": []
            }

        desdobles[bloque]["Materias"].append({
            "Asignatura": asignatura,
            "Grupo": grupo
        })

    print(f"{len(desdobles)} bloques de desdobles cargados correctamente.\n")
    return desdobles


def cargar_restricciones_temporales_excel():
    restricciones = []

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return restricciones

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "RestriccionesTemporales" not in wb.sheetnames:
        print("No existe la hoja 'RestriccionesTemporales' en el archivo Excel.\n")
        return restricciones

    ws = wb["RestriccionesTemporales"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    try:
        idx_tipo = encabezados.index("tipo")
        idx_asignatura = encabezados.index("asignatura")
        idx_grupo = encabezados.index("grupo")
        idx_dia = encabezados.index("dia")
        idx_sesion = encabezados.index("sesion")
        idx_valor = encabezados.index("valor")
    except ValueError:
        print("Error: La hoja 'RestriccionesTemporales' debe tener las columnas:")
        print("'Tipo', 'Asignatura', 'Grupo', 'Dia', 'Sesion', 'Valor'.\n")
        return restricciones

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        tipo = "" if row[idx_tipo] is None else str(row[idx_tipo]).strip()
        asignatura = "" if row[idx_asignatura] is None else str(row[idx_asignatura]).strip()
        grupo = "" if row[idx_grupo] is None else str(row[idx_grupo]).strip()
        dia = "" if row[idx_dia] is None else str(row[idx_dia]).strip()

        sesion = None
        if row[idx_sesion] is not None and str(row[idx_sesion]).strip() != "":
            try:
                sesion = int(row[idx_sesion])
            except ValueError:
                print(f"Restricción ignorada: sesión inválida en {tipo} - {asignatura} - {grupo}.")
                continue

        valor = "" if row[idx_valor] is None else str(row[idx_valor]).strip()

        restricciones.append({
            "Tipo": tipo,
            "Asignatura": asignatura,
            "Grupo": grupo,
            "Dia": dia,
            "Sesion": sesion,
            "Valor": valor
        })

    print(f"{len(restricciones)} restricciones temporales cargadas correctamente.\n")
    return restricciones


def convertir_sesion(valor, nombre_equipo, campo):
    if valor is None or str(valor).strip() == "":
        return None

    texto = str(valor).strip().replace(",", ".")

    try:
        return int(float(texto))
    except ValueError:
        raise ValueError(
            f"{campo} debe ser numérico en el equipo '{nombre_equipo}'. "
            f"Valor encontrado: {valor}"
        )


def cargar_equipos_excel():
    equipos = {}

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return equipos

    wb = load_workbook(ARCHIVO_EXCEL_DATOS, data_only=True)

    if "Equipos" not in wb.sheetnames:
        print("No existe la hoja 'Equipos' en el archivo Excel.\n")
        return equipos

    ws = wb["Equipos"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    try:
        idx_equipo = encabezados.index("equipo")
        idx_nombre = encabezados.index("nombre")
        idx_apellido1 = encabezados.index("apellido1")
        idx_apellido2 = encabezados.index("apellido2")
        idx_tipo = encabezados.index("tipo")
        idx_dia = encabezados.index("dia")
        idx_sesion_inicio = encabezados.index("sesioninicio")
        idx_sesion_fin = encabezados.index("sesionfin")
    except ValueError:
        print("Error: La hoja 'Equipos' debe tener las columnas:")
        print("'Equipo', 'Nombre', 'Apellido1', 'Apellido2', 'Tipo', 'Dia', 'SesionInicio', 'SesionFin'.\n")
        return equipos

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        equipo = "" if row[idx_equipo] is None else str(row[idx_equipo]).strip()

        # Si no hay equipo, ignoramos la fila.
        if equipo == "":
            continue

        nombre = "" if row[idx_nombre] is None else str(row[idx_nombre]).strip()
        apellido1 = "" if row[idx_apellido1] is None else str(row[idx_apellido1]).strip()
        apellido2 = "" if row[idx_apellido2] is None else str(row[idx_apellido2]).strip()
        tipo = "" if row[idx_tipo] is None else str(row[idx_tipo]).strip()
        dia = "" if row[idx_dia] is None else str(row[idx_dia]).strip()

        sesion_inicio = convertir_sesion(
            row[idx_sesion_inicio],
            equipo,
            "SesionInicio"
        )

        sesion_fin = convertir_sesion(
            row[idx_sesion_fin],
            equipo,
            "SesionFin"
        )

        tipo_norm = normalizar(tipo)

        if tipo_norm == "":
            tipo_norm = "preferente"

        if tipo_norm not in ["preferente", "obligatoria"]:
            raise ValueError(
                f"Tipo no válido en el equipo '{equipo}': {tipo}. "
                f"Debe ser 'Preferente' u 'Obligatoria'."
            )

        if tipo_norm == "obligatoria":
            if dia == "" or sesion_inicio is None or sesion_fin is None:
                raise ValueError(
                    f"El equipo obligatorio '{equipo}' debe tener Dia, "
                    f"SesionInicio y SesionFin."
                )

            if sesion_inicio > sesion_fin:
                raise ValueError(
                    f"En el equipo '{equipo}', SesionInicio no puede ser mayor que SesionFin."
                )

        if equipo not in equipos:
            equipos[equipo] = {
                "Tipo": tipo,
                "Dia": dia,
                "SesionInicio": sesion_inicio,
                "SesionFin": sesion_fin,
                "Miembros": []
            }

        equipos[equipo]["Miembros"].append({
            "Nombre": nombre,
            "Apellido1": apellido1,
            "Apellido2": apellido2
        })

    print(f"{len(equipos)} equipos cargados correctamente desde la hoja 'Equipos'.\n")
    return equipos


def cargar_disponibilidad_profesor_excel():
    disponibilidad = {}

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return disponibilidad

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "DisponibilidadProfesor" not in wb.sheetnames:
        print("No existe la hoja 'DisponibilidadProfesor' en el archivo Excel.\n")
        return disponibilidad

    ws = wb["DisponibilidadProfesor"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    try:
        idx_nombre = encabezados.index("nombre")
        idx_apellido1 = encabezados.index("apellido1")
        idx_apellido2 = encabezados.index("apellido2")
        idx_dia = encabezados.index("dia")
        idx_sesion = encabezados.index("sesion")
        idx_disponible = encabezados.index("disponible")
    except ValueError:
        print("Error: La hoja 'DisponibilidadProfesor' debe tener las columnas:")
        print("'Nombre', 'Apellido1', 'Apellido2', 'Dia', 'Sesion', 'Disponible'.\n")
        return disponibilidad

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        nombre = str(row[idx_nombre]).strip()
        apellido1 = str(row[idx_apellido1]).strip()
        apellido2 = str(row[idx_apellido2]).strip()
        dia = str(row[idx_dia]).strip()

        try:
            sesion = int(row[idx_sesion])
        except (TypeError, ValueError):
            print(f"Fila ignorada: sesión inválida para {nombre} {apellido1} {apellido2}.")
            continue

        disponible = str(row[idx_disponible]).strip().lower()
        clave_profesor = f"{nombre} {apellido1} {apellido2}"

        if clave_profesor not in disponibilidad:
            disponibilidad[clave_profesor] = {}

        if dia not in disponibilidad[clave_profesor]:
            disponibilidad[clave_profesor][dia] = {}

        disponibilidad[clave_profesor][dia][sesion] = disponible

    print(f"Disponibilidad de {len(disponibilidad)} profesores cargada correctamente.\n")
    return disponibilidad


def cargar_asignacion_docente_excel():
    asignaciones = []

    if not os.path.exists(ARCHIVO_ASIGNACION):
        print(f"No se encontró el archivo {ARCHIVO_ASIGNACION}.\n")
        return asignaciones

    wb = load_workbook(ARCHIVO_ASIGNACION, data_only=True)

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]

        encabezados = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in ws[1]
        ]

        if encabezados[:4] != ["clase", "asignatura", "horas", "profesor"]:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue

            try:
                asignaciones.append({
                    "Grupo": str(row[0]).strip(),
                    "Asignatura": str(row[1]).strip(),
                    "Horas": int(row[2]),
                    "Profesor": str(row[3]).strip()
                })
            except:
                continue

    print(f"{len(asignaciones)} asignaciones docentes cargadas correctamente.\n")
    return asignaciones


def cargar_max_sesiones_dia():
    max_sesiones = {}

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return max_sesiones

    wb = load_workbook(ARCHIVO_EXCEL_DATOS, data_only=True)

    if "Asignaturas" not in wb.sheetnames:
        print("No existe la hoja 'Asignaturas'.\n")
        return max_sesiones

    ws = wb["Asignaturas"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    try:
        idx_asignatura = encabezados.index("asignatura")
        idx_grupo = encabezados.index("grupo")
        idx_max = encabezados.index("maxsesionesdia")
    except ValueError:
        print("Error: La hoja 'Asignaturas' debe tener 'Asignatura', 'Grupo', 'MaxSesionesDia'.\n")
        return max_sesiones

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        asignatura = str(row[idx_asignatura]).strip()
        grupo = str(row[idx_grupo]).strip()

        try:
            max_dia = int(row[idx_max])
        except:
            max_dia = 1

        max_sesiones[(asignatura, grupo)] = max_dia

    return max_sesiones


# ------------------------------------------------------------
# UTILIDADES
# ------------------------------------------------------------

def clave_orden_clase(grupo):
    grupo = str(grupo).strip().upper()

    m = re.match(r"^(ESO|BACH)\s*(\d+)\s*([A-Z])$", grupo)

    if not m:
        return (999, 999, grupo)

    etapa = m.group(1)
    curso = int(m.group(2))
    letra = m.group(3)
    orden_etapa = 0 if etapa == "ESO" else 1

    return (orden_etapa, curso, letra)


def nombre_profesor(miembro):
    return f"{miembro['Nombre']} {miembro['Apellido1']} {miembro['Apellido2']}"


def profesor_disponible(profesor, dia, sesion, disponibilidad):
    if profesor not in disponibilidad:
        return True

    if dia not in disponibilidad[profesor]:
        return True

    if sesion not in disponibilidad[profesor][dia]:
        return True

    return disponibilidad[profesor][dia][sesion].lower() != "no"


def esta_libre_grupo(horario_grupos, grupo, dia, sesion):
    return (
        grupo not in horario_grupos or
        dia not in horario_grupos[grupo] or
        sesion not in horario_grupos[grupo][dia]
    )


def esta_libre_recurso(diccionario, recurso, dia, clave_hora):
    return (
        recurso not in diccionario or
        dia not in diccionario[recurso] or
        clave_hora not in diccionario[recurso][dia]
    )


def contar_asignatura_en_dia(horario_grupos, grupo, dia, asignatura):
    if grupo not in horario_grupos:
        return 0

    if dia not in horario_grupos[grupo]:
        return 0

    contador = 0

    for contenido in horario_grupos[grupo][dia].values():
        if isinstance(contenido, list):
            for datos in contenido:
                if datos["Asignatura"] == asignatura:
                    contador += 1
        else:
            if contenido["Asignatura"] == asignatura:
                contador += 1

    return contador


def obtener_aula(asignatura, grupo, dia, clave_hora, aulas, horario_aulas):
    config = None

    for a in aulas:
        if a["Asignatura"] == asignatura and a["Grupo"] == grupo:
            config = a
            break

    if config is None:
        return grupo

    tipo = config["TipoAula"].strip().lower()

    if tipo == "aulagrupo":
        return config["AulaFija"] if config["AulaFija"] else grupo

    if tipo == "aulaespecifica":
        return config["AulaFija"]

    if tipo == "aulaflexible":
        for aula in config["AulasPosibles"]:
            if esta_libre_recurso(horario_aulas, aula, dia, clave_hora):
                return aula
        return None

    return config["AulaFija"] if config["AulaFija"] else grupo


def restricciones_temporales_permiten(asignatura, grupo, dia, sesion, restricciones):
    for r in restricciones:
        if r["Asignatura"] != asignatura:
            continue

        if r["Grupo"] != grupo:
            continue

        tipo = r["Tipo"].strip().lower()

        if tipo == "nohora":
            if r["Dia"] == dia and r["Sesion"] == sesion:
                return False

        if tipo == "nodia":
            if r["Dia"] == dia:
                return False

        if tipo == "fijarhora":
            if not (r["Dia"] == dia and r["Sesion"] == sesion):
                return False

    return True


def obtener_duracion_bloque(asignatura, grupo, restricciones):
    for r in restricciones:
        if r["Asignatura"] == asignatura and r["Grupo"] == grupo:
            if r["Tipo"].strip().lower() == "dosseguidas":
                try:
                    valor = int(r["Valor"])
                    return max(2, valor)
                except:
                    return 2

    return 1

def tiene_restriccion_dos_seguidas(asignatura, grupo, restricciones):
    for r in restricciones:
        if r["Asignatura"] == asignatura and r["Grupo"] == grupo:
            if r["Tipo"].strip().lower() == "dosseguidas":
                return True
    return False

# ------------------------------------------------------------
# BLOQUEO DE REUNIONES
# ------------------------------------------------------------

def bloquear_reuniones_obligatorias(horario_profesores, equipos, horarios):
    for nombre_equipo, datos in equipos.items():
        if datos["Tipo"].strip().lower() != "obligatoria":
            continue

        dia = datos["Dia"]
        inicio = datos["SesionInicio"]
        fin = datos["SesionFin"]

        if not dia or inicio is None or fin is None:
            continue

        claves_hora = []

        # Se toma BACH como referencia para reuniones.
        for sesion in range(inicio, fin + 1):
            for franja in horarios["BACH"]:
                if franja["Sesion"] == sesion:
                    claves_hora.append(f"{franja['Inicio']}-{franja['Fin']}")

        for miembro in datos["Miembros"]:
            profesor = nombre_profesor(miembro)

            horario_profesores.setdefault(profesor, {})
            horario_profesores[profesor].setdefault(dia, {})

            for clave in claves_hora:
                horario_profesores[profesor][dia][clave] = {
                    "Asignatura": f"Reunión {nombre_equipo}",
                    "Grupo": "",
                    "Aula": "",
                    "Sesion": ""
                }


# ------------------------------------------------------------
# COLOCACIÓN DE BLOQUES
# ------------------------------------------------------------

def es_tutoria(asignatura):
    return asignatura.strip().lower() in ["tutoria", "tutoría"]


def preparar_bloques(asignaciones, desdobles, restricciones):
    bloques = []
    usadas_en_desdoble = set()

    # Índice rápido: (asignatura, grupo) -> asignación
    indice_asignaciones = {}

    for a in asignaciones:
        clave = (normalizar(a["Asignatura"]), normalizar(a["Grupo"]))
        indice_asignaciones[clave] = a

    # ------------------------------------------------------------
    # 1. BLOQUES DE DESDOBLE
    # ------------------------------------------------------------
    for nombre_bloque, datos_bloque in desdobles.items():
        tipo_desdoble = datos_bloque.get("Tipo", "G").strip().upper()
        materias = datos_bloque.get("Materias", [])

        if tipo_desdoble not in ["C", "G"]:
            tipo_desdoble = "G"

        elementos = []

        for m in materias:
            clave = (normalizar(m["Asignatura"]), normalizar(m["Grupo"]))

            if clave not in indice_asignaciones:
                print(f"Advertencia: {m['Asignatura']} - {m['Grupo']} del bloque {nombre_bloque} no está en asignacion_completa.xlsx.")
                continue

            a = indice_asignaciones[clave]
            elementos.append(a)
            usadas_en_desdoble.add(clave)

        if not elementos:
            continue

        horas = [e["Horas"] for e in elementos]

        if len(set(horas)) != 1:
            print(f"Error: el desdoble {nombre_bloque} tiene asignaturas con distinto número de horas.")
            for e in elementos:
                print(f" - {e['Asignatura']} | {e['Grupo']} | {e['Horas']} h")
            return []

        grupos_afectados = sorted(
            {e["Grupo"] for e in elementos},
            key=clave_orden_clase
        )

        for _ in range(horas[0]):
            bloques.append({
                "Tipo": "Desdoble",
                "TipoDesdoble": tipo_desdoble,
                "NombreBloque": nombre_bloque,
                "Duracion": 1,
                "Elementos": elementos,
                "GruposAfectados": grupos_afectados
            })

    # ------------------------------------------------------------
    # 2. BLOQUES NORMALES
    # ------------------------------------------------------------
    for a in asignaciones:
        clave = (normalizar(a["Asignatura"]), normalizar(a["Grupo"]))

        if clave in usadas_en_desdoble:
            continue

        duracion = obtener_duracion_bloque(
            a["Asignatura"],
            a["Grupo"],
            restricciones
        )

        horas = a["Horas"]

        if duracion > 1:
            bloques_dobles = horas // duracion
            horas_sueltas = horas % duracion

            for _ in range(bloques_dobles):
                bloques.append({
                    "Tipo": "Normal",
                    "TipoDesdoble": "",
                    "NombreBloque": "",
                    "Duracion": duracion,
                    "Elementos": [a],
                    "GruposAfectados": [a["Grupo"]]
                })

            for _ in range(horas_sueltas):
                bloques.append({
                    "Tipo": "Normal",
                    "TipoDesdoble": "",
                    "NombreBloque": "",
                    "Duracion": 1,
                    "Elementos": [a],
                    "GruposAfectados": [a["Grupo"]]
                })

        else:
            for _ in range(horas):
                bloques.append({
                    "Tipo": "Normal",
                    "TipoDesdoble": "",
                    "NombreBloque": "",
                    "Duracion": 1,
                    "Elementos": [a],
                    "GruposAfectados": [a["Grupo"]]
                })

    return bloques


def puede_colocar_bloque(bloque, dia, sesion_inicio, horarios, horario_grupos,
                         horario_profesores, horario_aulas, aulas,
                         restricciones, disponibilidad, max_sesiones):

    duracion = bloque["Duracion"]

    for desplazamiento in range(duracion):
        sesion = sesion_inicio + desplazamiento

        # ------------------------------------------------------------
        # 1. Comprobar que la sesión existe en TODOS los grupos afectados
        # ------------------------------------------------------------
        for grupo in bloque["GruposAfectados"]:
            sesiones_validas = [
                s["Sesion"]
                for s in obtener_sesiones_grupo_dia(grupo, dia, horarios)
            ]

            if sesion not in sesiones_validas:
                return False

        # ------------------------------------------------------------
        # 2. Comprobar grupo ocupado
        # ------------------------------------------------------------
        if bloque["Tipo"] == "Normal":
            grupo = bloque["Elementos"][0]["Grupo"]

            if not esta_libre_grupo(horario_grupos, grupo, dia, sesion):
                return False

        else:
            # En desdobles se reserva la franja en todos los grupos afectados.
            for grupo in bloque["GruposAfectados"]:
                if not esta_libre_grupo(horario_grupos, grupo, dia, sesion):
                    return False

        # ------------------------------------------------------------
        # 3. Comprobar profesores, aulas, restricciones y máximos diarios
        # ------------------------------------------------------------
        for elemento in bloque["Elementos"]:
            asignatura = elemento["Asignatura"]
            grupo = elemento["Grupo"]
            profesor = elemento["Profesor"]

            clave = clave_horaria(grupo, dia, sesion, horarios)

            if clave is None:
                return False

            if not esta_libre_recurso(horario_profesores, profesor, dia, clave):
                return False

            if not profesor_disponible(profesor, dia, sesion, disponibilidad):
                return False

            if not restricciones_temporales_permiten(asignatura, grupo, dia, sesion, restricciones):
                return False

            if bloque["Duracion"] == 1:
                if hay_misma_asignatura_consecutiva(horario_grupos, grupo, dia, sesion, asignatura):
                    return False

            max_dia = max_sesiones.get((asignatura, grupo), 1)

            if contar_asignatura_en_dia(horario_grupos, grupo, dia, asignatura) >= max_dia:
                return False

            aula = obtener_aula(asignatura, grupo, dia, clave, aulas, horario_aulas)

            if aula is None:
                return False

            if not esta_libre_recurso(horario_aulas, aula, dia, clave):
                return False

    return True


def colocar_sesion(horario_grupos, horario_profesores, horario_aulas,
                   grupo, profesor, aula, dia, sesion, asignatura, horarios):

    clave = clave_horaria(grupo, dia, sesion, horarios)

    if aula is None or str(aula).strip() == "":
        aula = "SIN AULA"

    horario_grupos.setdefault(grupo, {})
    horario_grupos[grupo].setdefault(dia, {})

    # En un grupo puede haber varias asignaturas a la vez si son desdobles.
    if sesion not in horario_grupos[grupo][dia]:
        horario_grupos[grupo][dia][sesion] = []

    horario_grupos[grupo][dia][sesion].append({
        "Asignatura": asignatura,
        "Grupo": grupo,
        "Profesor": profesor,
        "Aula": aula
    })

    horario_profesores.setdefault(profesor, {})
    horario_profesores[profesor].setdefault(dia, {})
    horario_profesores[profesor][dia][clave] = {
        "Asignatura": asignatura,
        "Grupo": grupo,
        "Aula": aula,
        "Sesion": sesion
    }

    horario_aulas.setdefault(aula, {})
    horario_aulas[aula].setdefault(dia, {})
    horario_aulas[aula][dia][clave] = {
        "Asignatura": asignatura,
        "Grupo": grupo,
        "Profesor": profesor,
        "Sesion": sesion
    }


def colocar_bloque(bloque, dia, sesion_inicio, horario_grupos,
                   horario_profesores, horario_aulas, aulas, horarios):

    duracion = bloque["Duracion"]

    for desplazamiento in range(duracion):
        sesion = sesion_inicio + desplazamiento

        # ------------------------------------------------------------
        # 1. Colocar en horario de profesores y aulas
        # ------------------------------------------------------------
        for elemento in bloque["Elementos"]:
            asignatura = elemento["Asignatura"]
            grupo_real = elemento["Grupo"]
            profesor = elemento["Profesor"]

            clave = clave_horaria(grupo_real, dia, sesion, horarios)
            aula = obtener_aula(asignatura, grupo_real, dia, clave, aulas, horario_aulas)

            if aula is None:
                aula = "SIN AULA"

            horario_profesores.setdefault(profesor, {})
            horario_profesores[profesor].setdefault(dia, {})
            horario_profesores[profesor][dia][clave] = {
                "Asignatura": asignatura,
                "Grupo": grupo_real,
                "Aula": aula,
                "Sesion": sesion
            }

            horario_aulas.setdefault(aula, {})
            horario_aulas[aula].setdefault(dia, {})
            horario_aulas[aula][dia][clave] = {
                "Asignatura": asignatura,
                "Grupo": grupo_real,
                "Profesor": profesor,
                "Sesion": sesion
            }

        # ------------------------------------------------------------
        # 2. Colocar en horario de grupos
        # ------------------------------------------------------------
        if bloque["Tipo"] == "Normal":
            elemento = bloque["Elementos"][0]

            asignatura = elemento["Asignatura"]
            grupo = elemento["Grupo"]
            profesor = elemento["Profesor"]

            clave = clave_horaria(grupo, dia, sesion, horarios)
            aula = obtener_aula(asignatura, grupo, dia, clave, aulas, horario_aulas)

            if aula is None:
                aula = "SIN AULA"

            horario_grupos.setdefault(grupo, {})
            horario_grupos[grupo].setdefault(dia, {})
            horario_grupos[grupo][dia][sesion] = [{
                "Asignatura": asignatura,
                "Grupo": grupo,
                "Profesor": profesor,
                "Aula": aula,
                "Bloque": ""
            }]

        else:
            tipo_desdoble = bloque.get("TipoDesdoble", "G")

            for grupo_destino in bloque["GruposAfectados"]:
                horario_grupos.setdefault(grupo_destino, {})
                horario_grupos[grupo_destino].setdefault(dia, {})

                elementos_visibles = []

                for elemento in bloque["Elementos"]:
                    asignatura = elemento["Asignatura"]
                    grupo_real = elemento["Grupo"]
                    profesor = elemento["Profesor"]

                    # Desdoble de curso:
                    # en cada clase aparecen todas las opciones del curso.
                    if tipo_desdoble == "C":
                        visible = True

                    # Desdoble de grupo:
                    # en cada clase solo aparecen las opciones de esa clase.
                    else:
                        visible = normalizar(grupo_real) == normalizar(grupo_destino)

                    if visible:
                        clave = clave_horaria(grupo_real, dia, sesion, horarios)
                        aula = obtener_aula(asignatura, grupo_real, dia, clave, aulas, horario_aulas)

                        if aula is None:
                            aula = "SIN AULA"

                        elementos_visibles.append({
                            "Asignatura": asignatura,
                            "Grupo": grupo_real,
                            "Profesor": profesor,
                            "Aula": aula,
                            "Bloque": bloque["NombreBloque"]
                        })

                # Aunque no haya elementos visibles, se reserva la franja.
                # Esto evita que otra asignatura se coloque en esa clase a la misma hora.
                horario_grupos[grupo_destino][dia][sesion] = elementos_visibles

def tipo_aula_asignatura(asignatura, grupo, aulas):
    for a in aulas:
        if a["Asignatura"] == asignatura and a["Grupo"]== grupo:
            return a["TipoAula"].strip().lower()
    return "aulagrupo"

def hay_misma_asignatura_consecutiva(horario_grupos, grupo, dia, sesion, asignatura):
    if grupo not in horario_grupos or dia not in horario_grupos[grupo]:
        return False

    for sesion_vecina in [sesion - 1, sesion + 1]:
        if sesion_vecina in horario_grupos[grupo][dia]:
            contenido = horario_grupos[grupo][dia][sesion_vecina]

            if isinstance(contenido, list):
                for item in contenido:
                    if item["Asignatura"] == asignatura:
                        return True
            else:
                if contenido["Asignatura"] == asignatura:
                    return True

    return False

def posiciones_validas_bloque(bloque, horarios, horario_grupos, horario_profesores,
                              horario_aulas, aulas, restricciones, disponibilidad, max_sesiones):
    posiciones = []

    for dia in DIAS:
        sesiones_ref = obtener_sesiones_grupo_dia(
            bloque["GruposAfectados"][0],
            dia,
            horarios
        )

        for franja in sesiones_ref:
            sesion_inicio = franja["Sesion"]

            if puede_colocar_bloque(
                bloque,
                dia,
                sesion_inicio,
                horarios,
                horario_grupos,
                horario_profesores,
                horario_aulas,
                aulas,
                restricciones,
                disponibilidad,
                max_sesiones
            ):
                posiciones.append((dia, sesion_inicio))

    random.shuffle(posiciones)
    return posiciones


def coste_posicion_bloque(bloque, dia, sesion_inicio, horario_grupos):
    coste = random.random()

    for elemento in bloque["Elementos"]:
        grupo = elemento["Grupo"]
        asignatura = elemento["Asignatura"]

        coste += contar_asignatura_en_dia(
            horario_grupos,
            grupo,
            dia,
            asignatura
        ) * 10

    return coste


def buscar_mejor_posicion_bloque(bloque, horarios, horario_grupos, horario_profesores,
                                 horario_aulas, aulas, restricciones, disponibilidad, max_sesiones):
    posiciones = posiciones_validas_bloque(
        bloque,
        horarios,
        horario_grupos,
        horario_profesores,
        horario_aulas,
        aulas,
        restricciones,
        disponibilidad,
        max_sesiones
    )

    if not posiciones:
        return None

    posiciones.sort(
        key=lambda pos: coste_posicion_bloque(
            bloque,
            pos[0],
            pos[1],
            horario_grupos
        )
    )

    mejores = posiciones[:min(3, len(posiciones))]
    return random.choice(mejores)


def elegir_bloque_mas_restrictivo(bloques_pendientes, horarios, horario_grupos, horario_profesores,
                                  horario_aulas, aulas, restricciones, disponibilidad, max_sesiones):
    mejor_bloque = None
    mejores_posiciones = None

    for bloque in bloques_pendientes:
        posiciones = posiciones_validas_bloque(
            bloque,
            horarios,
            horario_grupos,
            horario_profesores,
            horario_aulas,
            aulas,
            restricciones,
            disponibilidad,
            max_sesiones
        )

        if mejores_posiciones is None or len(posiciones) < len(mejores_posiciones):
            mejor_bloque = bloque
            mejores_posiciones = posiciones

        if len(posiciones) == 0:
            return mejor_bloque, mejores_posiciones

    return mejor_bloque, mejores_posiciones


def describir_bloque(bloque):
    texto = []
    for e in bloque["Elementos"]:
        texto.append(f"{e['Asignatura']} | {e['Grupo']} | {e['Profesor']}")
    return texto


def diagnosticar_bloques_iniciales(bloques, horarios, aulas, restricciones, disponibilidad, max_sesiones):
    horario_grupos = {}
    horario_profesores = {}
    horario_aulas = {}

    print("\n--- DIAGNÓSTICO INICIAL DE BLOQUES ---")

    sin_huecos = []

    for bloque in bloques:
        posiciones = posiciones_validas_bloque(
            bloque,
            horarios,
            horario_grupos,
            horario_profesores,
            horario_aulas,
            aulas,
            restricciones,
            disponibilidad,
            max_sesiones
        )

        if len(posiciones) == 0:
            sin_huecos.append(bloque)

    print(f"Bloques sin ningún hueco posible desde el inicio: {len(sin_huecos)}")

    for bloque in sin_huecos[:30]:
        for e in bloque["Elementos"]:
            print(f" - {e['Asignatura']} | {e['Grupo']} | {e['Profesor']}")

    print()


def horas_semanales_grupo(grupo, horarios):
    total = 0

    for dia in DIAS:
        total += len(obtener_sesiones_grupo_dia(grupo, dia, horarios))

    return total


def validar_horas_por_grupo(asignaciones, horarios, desdobles):
    horas_por_grupo = {}
    detalle_por_grupo = {}
    asignaturas_en_desdoble = set()

    # 1. Contar desdobles
    for nombre_bloque, datos_bloque in desdobles.items():

        # Formato nuevo:
        # desdobles[bloque] = {"Tipo": "C/G", "Materias": [...]}
        if isinstance(datos_bloque, dict):
            tipo_desdoble = datos_bloque.get("Tipo", "G")
            materias = datos_bloque.get("Materias", [])
        else:
            # Compatibilidad con formato antiguo
            tipo_desdoble = "G"
            materias = datos_bloque

        materias_por_grupo = {}

        for m in materias:
            grupo = m["Grupo"]
            asignatura = m["Asignatura"]

            if grupo not in materias_por_grupo:
                materias_por_grupo[grupo] = []

            materias_por_grupo[grupo].append(asignatura)

        for grupo, lista_asignaturas in materias_por_grupo.items():
            horas_bloque = None

            for a in asignaciones:
                if a["Grupo"] == grupo and a["Asignatura"] in lista_asignaturas:
                    horas_bloque = a["Horas"]
                    asignaturas_en_desdoble.add((a["Asignatura"], a["Grupo"]))

            if horas_bloque is not None:
                horas_por_grupo[grupo] = horas_por_grupo.get(grupo, 0) + horas_bloque

                if grupo not in detalle_por_grupo:
                    detalle_por_grupo[grupo] = []

                detalle_por_grupo[grupo].append(
                    f"Desdoble {nombre_bloque} ({tipo_desdoble}): {horas_bloque} h"
                )

    # 2. Contar asignaturas normales
    for a in asignaciones:
        clave = (a["Asignatura"], a["Grupo"])

        if clave in asignaturas_en_desdoble:
            continue

        grupo = a["Grupo"]
        horas_por_grupo[grupo] = horas_por_grupo.get(grupo, 0) + a["Horas"]

        if grupo not in detalle_por_grupo:
            detalle_por_grupo[grupo] = []

        detalle_por_grupo[grupo].append(
            f"{a['Asignatura']}: {a['Horas']} h"
        )

    errores = []

    for grupo, horas_asignadas in sorted(horas_por_grupo.items(), key=lambda x: clave_orden_clase(x[0])):
        horas_necesarias = horas_semanales_grupo(grupo, horarios)

        if horas_asignadas != horas_necesarias:
            errores.append((grupo, horas_asignadas, horas_necesarias))

    if errores:
        print("\nError: hay grupos cuya carga horaria no coincide con las franjas semanales.\n")

        for grupo, horas_asignadas, horas_necesarias in errores:
            print(f"=== {grupo} ===")
            print(f"Horas configuradas: {horas_asignadas}")
            print(f"Franjas necesarias: {horas_necesarias}")
            print(f"Diferencia: {horas_asignadas - horas_necesarias}")
            print("Detalle:")

            for linea in detalle_por_grupo.get(grupo, []):
                print(f"  - {linea}")

            print()

        return False

    print("Validación correcta: todos los grupos tienen las horas necesarias para completar la semana.\n")
    return True


# ------------------------------------------------------------
# GENERACIÓN DEL HORARIO
# ------------------------------------------------------------

def generar_horarios(intentos=100):
    horarios = cargar_franjas_horarias()
    asignaciones = cargar_asignacion_docente_excel()
    aulas = cargar_aulas_excel()
    desdobles = cargar_desdobles_excel()
    restricciones = cargar_restricciones_temporales_excel()
    equipos = cargar_equipos_excel()
    disponibilidad = cargar_disponibilidad_profesor_excel()
    max_sesiones = cargar_max_sesiones_dia()

    if not asignaciones:
        print("No hay asignaciones docentes cargadas.\n")
        return None

    grupos = obtener_grupos_asignaciones(asignaciones)

    if not validar_horas_por_grupo(asignaciones, horarios, desdobles):
        print("La validación de horas ha fallado. Se intentará generar igualmente, pero puede haber huecos o exceso de horas.\n")

    bloques_base = preparar_bloques(asignaciones, desdobles, restricciones)

    if not bloques_base:
        print("No se han podido preparar los bloques de horario.\n")
        return None

    print(f"Se han generado {len(bloques_base)} bloques de colocación.\n")

    mejor_resultado = None
    mejor_colocados = -1
    mejor_incidencias = []

    for intento in range(1, intentos + 1):
        horario_grupos = {}
        horario_profesores = {}
        horario_aulas = {}
        incidencias_totales = []

        bloquear_reuniones_obligatorias(horario_profesores, equipos, horarios)

        bloques_pendientes = bloques_base[:]

        random.shuffle(bloques_pendientes)

        bloques_pendientes.sort(
            key=lambda b: (
                0 if b["Tipo"] == "Desdoble" else 1,
                -len(b["Elementos"]),
                -b["Duracion"],
                random.random()
            )
        )

        colocados = 0
        fallo = None

        for n, bloque in enumerate(bloques_pendientes, start=1):
            mejor = buscar_mejor_posicion_bloque(
                    bloque,
                    horarios,
                    horario_grupos,
                    horario_profesores,
                    horario_aulas,
                    aulas,
                    restricciones,
                    disponibilidad,
                    max_sesiones
                )

            if mejor is None:
                fallo = bloque
                incidencias_totales.append({
                    "Bloque": describir_bloque(bloque),
                    "Incidencia ": "No existe ninguna franja válida para este bloque"
                })
                break

            dia, sesion_inicio = mejor

            colocar_bloque(
                bloque,
                dia,
                sesion_inicio,
                horario_grupos,
                horario_profesores,
                horario_aulas,
                aulas,
                horarios
            )

            colocados += 1

        if colocados > mejor_colocados:
            mejor_colocados = colocados
            mejor_resultado = (horario_grupos, horario_profesores, horario_aulas)
            mejor_incidencias = incidencias_totales[:]

        if colocados == len(bloques_pendientes):
            print(f"Horario completo generado correctamente en el intento {intento}.\n")

            exportar_horarios_excel_con_incidencias(
                horario_grupos,
                horario_profesores,
                horario_aulas,
                horarios,
                incidencias_totales,
                grupos
            )

            print(f"Horario generado en {ARCHIVO_HORARIOS}.")
            print(f"Incidencias registradas: {len(incidencias_totales)}\n")

            return horario_grupos, horario_profesores, horario_aulas

        if intento % 10 == 0:
            print(f"Intento {intento}/{intentos} | Mejor resultado: {mejor_colocados}/{len(bloques_pendientes)} bloques colocados")

    print("No se ha podido generar un horario completo.")
    print(f"Mejor resultado: {mejor_colocados}/{len(bloques_base)} bloques colocados.")
    print("Se exporta el mejor horario parcial para revisar.\n")

    if mejor_resultado is not None:
        horario_grupos, horario_profesores, horario_aulas = mejor_resultado

        exportar_horarios_excel_con_incidencias(
            horario_grupos,
            horario_profesores,
            horario_aulas,
            horarios,
            mejor_incidencias,
            grupos
        )

    return None


# ------------------------------------------------------------
# EXPORTACIÓN
# ------------------------------------------------------------

def limpiar_nombre_hoja(nombre):
    invalidos = ['\\', '/', '*', '?', ':', '[', ']']

    for c in invalidos:
        nombre = nombre.replace(c, "-")

    return str(nombre)[:31]


def escribir_hoja_horario(ws, datos, horarios, grupo_hoja=None):
    ws.append(["Día", "Sesión", "Hora", "Asignatura", "Grupo", "Profesor", "Aula", "Bloque"])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --------------------------------------------------
    # HORARIO DE GRUPO: muestra TODAS las franjas
    # --------------------------------------------------
    if grupo_hoja is not None:
        for dia in DIAS:
            sesiones = obtener_sesiones_grupo_dia(grupo_hoja, dia, horarios)

            for franja in sesiones:
                sesion = franja["Sesion"]
                hora = f"{franja['Inicio']}-{franja['Fin']}"

                contenido = datos.get(dia, {}).get(sesion, None)

                if contenido is None:
                    ws.append([dia, sesion, hora, "", grupo_hoja, "", "", ""])
                    continue

                if not contenido:
                    ws.append([dia, sesion, hora, "RESERVADO DESDOBLE", grupo_hoja, "", "", ""])
                    continue

                for item in contenido:
                    ws.append([
                        dia,
                        sesion,
                        hora,
                        item.get("Asignatura", ""),
                        item.get("Grupo", ""),
                        item.get("Profesor", ""),
                        item.get("Aula", ""),
                        item.get("Bloque", "")
                    ])

    # --------------------------------------------------
    # HORARIO DE PROFESOR/AULA: solo sesiones ocupadas
    # --------------------------------------------------
    else:
        for dia in DIAS:
            if dia not in datos:
                continue

            for clave in sorted(datos[dia].keys(), key=lambda x: str(x)):
                contenido = datos[dia][clave]

                ws.append([
                    dia,
                    contenido.get("Sesion", ""),
                    clave,
                    contenido.get("Asignatura", ""),
                    contenido.get("Grupo", ""),
                    contenido.get("Profesor", ""),
                    contenido.get("Aula", ""),
                    ""
                ])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 18


def exportar_horarios_excel(horario_grupos, horario_profesores, horario_aulas, horarios):
    exportar_horarios_excel_con_incidencias(
        horario_grupos,
        horario_profesores,
        horario_aulas,
        horarios,
        [],
        sorted(horario_grupos.keys(), key=clave_orden_clase)
    )

def exportar_horarios_excel_con_incidencias(horario_grupos, horario_profesores,
                                            horario_aulas, horarios, incidencias,
                                            grupos=None):
    wb = Workbook()
    ws_inicial = wb.active
    wb.remove(ws_inicial)

    if grupos is None:
        grupos = sorted(horario_grupos.keys(), key=clave_orden_clase)

    # Hojas por grupo: se crean aunque el grupo tenga huecos o esté vacío
    for grupo in sorted(grupos, key=clave_orden_clase):
        ws = wb.create_sheet(limpiar_nombre_hoja(grupo))
        escribir_hoja_horario(
            ws,
            horario_grupos.get(grupo, {}),
            horarios,
            grupo_hoja=grupo
        )

    # Hojas por profesor
    for profesor in sorted(horario_profesores.keys()):
        nombre_hoja = limpiar_nombre_hoja(profesor)
        base = nombre_hoja
        n = 1

        while nombre_hoja in wb.sheetnames:
            nombre_hoja = limpiar_nombre_hoja(f"{base}_{n}")
            n += 1

        ws = wb.create_sheet(nombre_hoja)
        escribir_hoja_horario(ws, horario_profesores[profesor], horarios)

    # Hojas por aula
    for aula in sorted(horario_aulas.keys(), key=lambda x: str(x)):
        nombre_hoja = limpiar_nombre_hoja(f"Aula {aula}")
        base = nombre_hoja
        n = 1

        while nombre_hoja in wb.sheetnames:
            nombre_hoja = limpiar_nombre_hoja(f"{base}_{n}")
            n += 1

        ws = wb.create_sheet(nombre_hoja)
        escribir_hoja_horario(ws, horario_aulas[aula], horarios)

    # Hoja de incidencias
    ws_inc = wb.create_sheet("Incidencias")
    ws_inc.append(["Bloque", "Incidencia"])

    for cell in ws_inc[1]:
        cell.font = Font(bold=True)

    for inc in incidencias:
    # Puede ocurrir que alguna incidencia no tenga exactamente la clave "Incidencia".
    # Por eso se usa get() para evitar que el programa se rompa al exportar.
    
        bloque = inc.get("Bloque", "")

        if isinstance(bloque, list):
            bloque = " / ".join(str(x) for x in bloque)
        else:
            bloque = str(bloque)

        texto_incidencia = inc.get("Incidencia", "")

        # Compatibilidad por si en algún punto se hubiera usado otra clave
        if texto_incidencia == "":
            texto_incidencia = inc.get("incidencia", "")

        if texto_incidencia == "":
            texto_incidencia = inc.get("Motivo", "")

        if texto_incidencia == "":
            texto_incidencia = "Incidencia no especificada"

        ws_inc.append([bloque, texto_incidencia])

    ws_inc.column_dimensions["A"].width = 80
    ws_inc.column_dimensions["B"].width = 120

    wb.save(ARCHIVO_HORARIOS)
    print(f"Archivo {ARCHIVO_HORARIOS} generado correctamente.\n")
