# ============================================================
# AUTORÍA DEL CÓDIGO
# Autor: Aitor Santamaria Zuluaga
# Grado en Ingeniería en Organización Industrial
# JULIO DEL 2026
# ============================================================

# ============================================================
# 1. LIBRERÍAS
# ============================================================

import pandas as pd
from ortools.sat.python import cp_model
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


# ============================================================
# 2. CONFIGURACIÓN GENERAL
# ============================================================

ARCHIVO_DATOS = "datos_centro.xlsx"
ARCHIVO_ASIGNACION = "asignacion_completa.xlsx"
ARCHIVO_SALIDA = "HorarioFinal.xlsx"


# ============================================================
# 3. FUNCIONES DE NORMALIZACIÓN DE DATOS
# ============================================================

def normalizar_grupo(g):
    """
    Convierte grupos tipo 'BACH 1A', 'BACH1A', 'ESO 3AB' en una forma comparable.
    """
    return str(g).strip().upper().replace(" ", "")

def formatear_grupo(g_norm):
    """
    Devuelve el grupo con espacio: BACH1A -> BACH 1A, ESO3B -> ESO 3B.
    """
    g_norm = normalizar_grupo(g_norm)

    if g_norm.startswith("BACH"):
        return "BACH " + g_norm[4:]
    elif g_norm.startswith("ESO"):
        return "ESO " + g_norm[3:]
    else:
        return g_norm

def expandir_grupo(g):
    """
    Si recibe BACH 1AB, devuelve ['BACH 1A', 'BACH 1B'].
    Si recibe ESO 3A, devuelve ['ESO 3A'].
    """
    g_norm = normalizar_grupo(g)

    if g_norm.endswith("AB"):
        raiz = g_norm[:-2]
        return [formatear_grupo(raiz + "A"), formatear_grupo(raiz + "B")]

    return [formatear_grupo(g_norm)]

def pertenece_grupo(grupo_fila, grupo_objetivo):
    """
    Comprueba si una fila pertenece a un grupo individual.
    Ejemplo:
    grupo_fila = BACH 1AB
    grupo_objetivo = BACH 1A
    devuelve True.
    """
    objetivo = normalizar_grupo(grupo_objetivo)
    expandidos = [normalizar_grupo(g) for g in expandir_grupo(grupo_fila)]
    return objetivo in expandidos

def normalizar_dia(dia):
    """
    Convierte el nombre del día en índice numérico.
    """
    dia = str(dia).strip().lower()

    mapa = {
        "lunes": 0,
        "martes": 1,
        "miércoles": 2,
        "miercoles": 2,
        "jueves": 3,
        "viernes": 4
    }

    if dia not in mapa:
        raise ValueError(f"Día no válido en la hoja Equipos: {dia}")

    return mapa[dia]

def convertir_sesion(valor, nombre_equipo, campo):
    if valor is None:
        return None

    texto = str(valor).strip()

    if texto == "":
        return None

    texto = texto.replace(",", ".")

    try:
        return int(float(texto))
    except:
        raise ValueError(
            f"{campo} debe ser numérico en el equipo '{nombre_equipo}'. "
            f"Valor encontrado: '{valor}'"
        )

def cargar_equipos(xl_datos):
    """
    Carga la hoja 'Equipos' de datos_centro.xlsx.

    Estructura esperada:
    Equipo | Nombre | Apellido1 | Apellido2 | Tipo | Dia | SesionInicio | SesionFin

    Tipo:
    - Obligatoria: restricción dura. Debe indicarse día, sesión inicial y sesión final.
    - Preferente: restricción blanda. El modelo intentará buscar una hora común semanal.
    """

    df_equipos = pd.read_excel(xl_datos, sheet_name="Equipos")
    df_equipos = df_equipos.fillna("")
    df_equipos.columns = df_equipos.columns.astype(str).str.strip()

    columnas_obligatorias = [
        "Equipo", "Nombre", "Apellido1", "Apellido2",
        "Tipo", "Dia", "SesionInicio", "SesionFin"
    ]

    for col in columnas_obligatorias:
        if col not in df_equipos.columns:
            raise ValueError(
                f"Falta la columna obligatoria '{col}' en la hoja Equipos."
            )

    equipos_dict = {}

    for _, fila in df_equipos.iterrows():
        equipo = str(fila["Equipo"]).strip()

        # Si no hay equipo, no se considera una reunión de equipo.
        if equipo == "":
            continue

        nombre = str(fila["Nombre"]).strip()
        apellido1 = str(fila["Apellido1"]).strip()
        apellido2 = str(fila["Apellido2"]).strip()

        profesor = f"{nombre} {apellido1} {apellido2}".strip()

        if profesor == "":
            continue

        tipo = str(fila["Tipo"]).strip()
        dia = str(fila["Dia"]).strip()
        sesion_inicio = str(fila["SesionInicio"]).strip()
        sesion_fin = str(fila["SesionFin"]).strip()

        if equipo not in equipos_dict:
            equipos_dict[equipo] = {
                "equipo": equipo,
                "miembros": [],
                "tipo": "",
                "dia": "",
                "sesion_inicio": "",
                "sesion_fin": ""
            }

        equipos_dict[equipo]["miembros"].append(profesor)

        if tipo != "":
            equipos_dict[equipo]["tipo"] = tipo

        if dia != "":
            equipos_dict[equipo]["dia"] = dia

        if sesion_inicio != "":
            equipos_dict[equipo]["sesion_inicio"] = sesion_inicio

        if sesion_fin != "":
            equipos_dict[equipo]["sesion_fin"] = sesion_fin

    equipos = []

    for equipo, datos in equipos_dict.items():
        datos["miembros"] = sorted(list(set(datos["miembros"])))

        tipo = datos["tipo"].strip().lower()

        if tipo == "":
            tipo = "preferente"

        if tipo not in ["preferente", "obligatoria"]:
            raise ValueError(
                f"Tipo de equipo no válido en '{equipo}': {datos['tipo']}. "
                f"Debe ser 'Preferente' u 'Obligatoria'."
            )

        datos["tipo"] = tipo

        if tipo == "obligatoria":
            if datos["dia"] == "" or datos["sesion_inicio"] == "" or datos["sesion_fin"] == "":
                raise ValueError(
                    f"El equipo obligatorio '{equipo}' debe tener Dia, "
                    f"SesionInicio y SesionFin."
                )

            datos["dia_idx"] = normalizar_dia(datos["dia"])

            try:
                datos["sesion_inicio"] = convertir_sesion(
                    datos["sesion_inicio"],
                    equipo,
                    "SesionInicio"
                )
                datos["sesion_fin"] = convertir_sesion(
                    datos["sesion_fin"],
                    equipo,
                    "SesionFin"
                )
            except ValueError as e:
                raise ValueError(str(e))

        equipos.append(datos)

    return equipos


# ============================================================
# 4. FUNCIONES DE CARGA DE DATOS
# ============================================================


def cargar_afinidades(xl_datos):
    """
    Carga la pestaña 'Afinidades' del archivo datos_centro.xlsx.

    Estructura esperada:

                BACH1AB     BACH2AB     ESO3A
    Nombre      Biologia    Biologia    Biologia
    Jon         0           0           1
    ...

    Columnas fijas:
    Nombre | Apellido1 | Apellido2

    Códigos:
    0 = el profesor no es afín para dar la asignatura en ese grupo
    1 = el profesor puede impartir la asignatura en ese grupo
    2 = el profesor debe impartir la asignatura en ese grupo
    """

    df_raw = pd.read_excel(xl_datos, sheet_name="Afinidades", header=None)
    df_raw = df_raw.fillna("")

    # La fila 0 contiene los grupos: BACH1AB, BACH2AB, ESO3A...
    # La fila 1 contiene las asignaturas: Biologia, Biologia...
    fila_grupos = df_raw.iloc[0]
    fila_asignaturas = df_raw.iloc[1]

    afinidades = {}
    profesores = []

    # Los datos de profesores empiezan en la fila 2
    for idx in range(2, len(df_raw)):
        nombre = str(df_raw.iloc[idx, 0]).strip()
        apellido1 = str(df_raw.iloc[idx, 1]).strip()
        apellido2 = str(df_raw.iloc[idx, 2]).strip()

        profesor = f"{nombre} {apellido1} {apellido2}".strip()

        if profesor == "":
            continue

        profesores.append(profesor)

        # Las afinidades empiezan en la columna 3
        for col in range(3, len(df_raw.columns)):
            grupo = str(fila_grupos[col]).strip()
            asignatura = str(fila_asignaturas[col]).strip()

            if grupo == "" or asignatura == "":
                continue

            grupo_norm = normalizar_grupo(grupo)
            asignatura_limpia = asignatura.strip()

            try:
                valor = int(df_raw.iloc[idx, col])
            except:
                valor = 0

            afinidades[(profesor, asignatura_limpia, grupo_norm)] = valor

    profesores = sorted(list(set(profesores)))

    return afinidades, profesores



def cargar_max_horas_profesores(xl_datos):
    """
    Carga la hoja 'Profesores' de datos_centro.xlsx.

    Estructura esperada:
    Nombre | Apellido1 | Apellido2 | MaxHoras

    Devuelve un diccionario:
    {
        "Nombre Apellido1 Apellido2": MaxHoras
    }
    """

    df_prof = pd.read_excel(xl_datos, sheet_name="Profesores")
    df_prof = df_prof.fillna("")

    df_prof.columns = df_prof.columns.astype(str).str.strip()

    columnas_obligatorias = ["Nombre", "Apellido1", "Apellido2", "MaxHoras"]

    for col in columnas_obligatorias:
        if col not in df_prof.columns:
            raise ValueError(
                f"Falta la columna obligatoria '{col}' en la hoja Profesores."
            )

    max_horas_profesor = {}

    for _, fila in df_prof.iterrows():
        nombre = str(fila["Nombre"]).strip()
        apellido1 = str(fila["Apellido1"]).strip()
        apellido2 = str(fila["Apellido2"]).strip()

        profesor = f"{nombre} {apellido1} {apellido2}".strip()

        if profesor == "":
            continue

        try:
            max_horas = int(fila["MaxHoras"])
        except:
            raise ValueError(
                f"El profesor {profesor} no tiene un valor válido en MaxHoras."
            )

        max_horas_profesor[profesor] = max_horas

    return max_horas_profesor



# ============================================================
# 5. FUNCIONES AUXILIARES SOBRE ASIGNATURAS Y GRUPOS
# ============================================================

def es_tutoria(nombre_asignatura):
    """
    Detecta si una asignatura corresponde a Tutoría.
    """
    return str(nombre_asignatura).strip().lower() in [
        "tutoria",
        "tutoría"
    ]


def filas_tutoria_del_grupo(df_total, grupo_objetivo):
    """
    Devuelve las filas de Tutoría que pertenecen al grupo indicado.
    """
    return [
        i for i in df_total.index
        if pertenece_grupo(df_total.loc[i, 'Grupo'], grupo_objetivo)
        and es_tutoria(df_total.loc[i, 'Asignatura'])
    ]

def filas_no_tutoria_del_grupo(df_total, grupo_objetivo):
    """
    Devuelve las filas de asignaturas que pertenecen al grupo indicado
    y que no son Tutoría.
    """
    return [
        i for i in df_total.index
        if pertenece_grupo(df_total.loc[i, 'Grupo'], grupo_objetivo)
        and not es_tutoria(df_total.loc[i, 'Asignatura'])
    ]


def actividades_representantes_de_grupo(df_total, grupo_objetivo):
    """
    Devuelve las filas que cuentan como actividades lectivas para un grupo.

    Las asignaturas normales cuentan una a una.
    Las asignaturas con el mismo BloqueDesdoble cuentan una sola vez,
    usando una fila representante del bloque.

    Ejemplo:
    ESO 3A:
        Matemáticas                 -> cuenta 1
        Desdoble1 - Biología         -> cuenta 1
        Desdoble1 - Física y Química -> NO suma otra vez, va sincronizada
    """

    idxs_g = [
        i for i in df_total.index
        if pertenece_grupo(df_total.loc[i, 'Grupo'], grupo_objetivo)
    ]

    representantes = []

    # Asignaturas sin desdoble
    for i in idxs_g:
        if str(df_total.loc[i, 'BloqueDesdoble']).strip() == "":
            representantes.append(i)

    # Asignaturas con desdoble: un representante por bloque
    bloques_vistos = set()

    for i in idxs_g:
        bloque = str(df_total.loc[i, 'BloqueDesdoble']).strip()

        if bloque == "":
            continue

        clave_bloque = (
            normalizar_grupo(df_total.loc[i, 'Grupo']),
            bloque
        )

        if clave_bloque not in bloques_vistos:
            representantes.append(i)
            bloques_vistos.add(clave_bloque)

    return representantes


# ============================================================
# 6. FUNCIONES AUXILIARES SOBRE HORARIOS
# ============================================================


def slot_permitido(grupo, d, s):
    """
    Indica si un grupo puede tener clase en un día y slot determinado.
    d: 0 lunes, ..., 4 viernes
    s: 0 a 7
    """
    grupo = normalizar_grupo(grupo)

    es_bach = "BACH" in grupo
    es_1_bach = es_bach and "1" in grupo
    es_2_bach = es_bach and "2" in grupo

    if es_bach:
        if s == 7:
            return False

        if es_2_bach and d in [3, 4] and s == 6:
            return False

        if es_1_bach and d == 4 and s == 6:
            return False

        return True

    else:
        if s in [0, 5]:
            return False

        return True



def pares_consecutivos_permitidos(grupo, d):
    """
    Pares de slots que pueden contar como dos horas seguidas.
    No cruza recreos ni comidas.
    """
    grupo = normalizar_grupo(grupo)

    if "BACH" in grupo:
        posibles = [(0, 1), (1, 2), (3, 4), (4, 5), (5, 6)]
    else:
        posibles = [(1, 2), (3, 4), (6, 7)]

    return [(s1, s2) for s1, s2 in posibles if slot_permitido(grupo, d, s1) and slot_permitido(grupo, d, s2)]


def aplicar_restricciones_equipos(model, equipos, ocup_prof, profesores_disponibles, dias, slots):
    """
    Aplica las restricciones de equipos.

    Equipos obligatorios:
    - Restricción dura.
    - Todos los miembros deben estar libres en el día y sesiones indicadas.

    Equipos preferentes:
    - Restricción blanda.
    - El modelo intenta encontrar una sesión semanal común libre para todos los miembros.
    """

    penalizaciones_equipos = []

    PESO_EQUIPO_PREFERENTE_NO_ASIGNADO = 300

    for equipo in equipos:
        nombre_equipo = equipo["equipo"]
        miembros = [
            p for p in equipo["miembros"]
            if p in profesores_disponibles
        ]

        if not miembros:
            print(
                f"[AVISO] El equipo '{nombre_equipo}' no tiene miembros reconocidos "
                f"en la lista de profesores disponibles."
            )
            continue

        tipo = equipo["tipo"]

        # -----------------------------------------------------
        # EQUIPOS OBLIGATORIOS
        # -----------------------------------------------------
        if tipo == "obligatoria":
            d = equipo["dia_idx"]
            inicio = equipo["sesion_inicio"]
            fin = equipo["sesion_fin"]

            # Interpretación: SesionInicio y SesionFin son inclusivas.
            # Ejemplo: inicio=1, fin=2 bloquea las sesiones 1 y 2.
            for s in range(inicio, fin + 1):
                if s not in slots:
                    raise ValueError(
                        f"El equipo obligatorio '{nombre_equipo}' usa la sesión {s}, "
                        f"pero esa sesión no existe en el modelo."
                    )

                for p in miembros:
                    model.Add(ocup_prof[(p, d, s)] == 0)

        # -----------------------------------------------------
        # EQUIPOS PREFERENTES
        # -----------------------------------------------------
        elif tipo == "preferente":
            posibles_reuniones = []

            for d in dias:
                for s in slots:
                    r = model.NewBoolVar(
                        f"reunion_pref_{nombre_equipo}_{d}_{s}".replace(" ", "_").replace("/", "_")
                    )

                    # Si r = 1, todos los miembros deben estar libres.
                    for p in miembros:
                        model.Add(r <= 1 - ocup_prof[(p, d, s)])

                    posibles_reuniones.append(r)

            if posibles_reuniones:
                # Como máximo se asigna una sesión común para ese equipo.
                model.Add(sum(posibles_reuniones) <= 1)

                falta_reunion = model.NewBoolVar(
                    f"falta_reunion_{nombre_equipo}".replace(" ", "_").replace("/", "_")
                )

                # Si no se puede asignar reunión, falta_reunion = 1.
                model.Add(sum(posibles_reuniones) + falta_reunion == 1)

                penalizaciones_equipos.append(
                    PESO_EQUIPO_PREFERENTE_NO_ASIGNADO * falta_reunion
                )

    return penalizaciones_equipos



# ============================================================
# 7. FUNCIONES DE VALIDACIÓN Y DIAGNÓSTICO
# ============================================================


def comprobar_unico_profesor_obligatorio(df_total, afinidades, profesores_disponibles):
    """
    Comprueba que cada asignatura-grupo tenga, como máximo,
    un único profesor con afinidad 2.

    Si hay más de un profesor con valor 2 para la misma asignatura-grupo,
    se lanza un error, porque la preferencia obligatoria no sería unívoca.
    """

    errores = []

    for _, fila in df_total.iterrows():
        asignatura = str(fila['Asignatura']).strip()
        grupo = str(fila['Grupo']).strip()
        grupo_norm = normalizar_grupo(grupo)

        profesores_con_2 = []

        for p in profesores_disponibles:
            valor = afinidades.get((p, asignatura, grupo_norm), 0)

            if valor == 2:
                profesores_con_2.append(p)

        if len(profesores_con_2) > 1:
            errores.append(
                f"{asignatura} - {grupo}: {profesores_con_2}"
            )

    if errores:
        mensaje = (
            "Hay asignaturas con más de un profesor obligatorio, es decir, "
            "con más de un valor 2 en la tabla de afinidades:\n\n"
        )

        mensaje += "\n".join(f"  - {e}" for e in errores)

        raise ValueError(mensaje)



def diagnosticar_posibles_bloqueos(df_total, afinidades, profesores_disponibles):
    """
    Diagnóstico básico para casos en los que las horas cuadran,
    pero el modelo CP-SAT sigue siendo inviable.
    """

    print("\n--- DIAGNÓSTICO DE POSIBLES BLOQUEOS ---")

    # 1. Asignaturas con pocos profesores posibles
    print("\n1) Asignaturas con pocos profesores posibles:")

    for i, fila in df_total.iterrows():
        asignatura = str(fila['Asignatura']).strip()
        grupo_norm = normalizar_grupo(fila['Grupo'])
        grupo = str(fila['Grupo']).strip()

        obligatorios = []
        afines = []

        for p in profesores_disponibles:
            valor = afinidades.get((p, asignatura, grupo_norm), 0)

            if valor == 2:
                obligatorios.append(p)
            elif valor == 1:
                afines.append(p)

        if obligatorios:
            posibles = obligatorios
            tipo = "OBLIGATORIOS"
        else:
            posibles = afines
            tipo = "AFINES"

        if len(posibles) <= 1:
            print(
                f"  [!] {asignatura} - {grupo}: "
                f"{len(posibles)} profesor posible ({tipo}) -> {posibles}"
            )

    # 2. Desdobles con un único profesor posible común
    print("\n2) Revisión de desdobles:")

    bloques_dict = {}

    for i, fila in df_total.iterrows():
        bloque = str(fila['BloqueDesdoble']).strip()

        if bloque == "":
            continue

        clave = (normalizar_grupo(fila['Grupo']), bloque)

        if clave not in bloques_dict:
            bloques_dict[clave] = []

        bloques_dict[clave].append(i)

    for clave, idxs in bloques_dict.items():
        grupo_norm, bloque = clave

        if len(idxs) <= 1:
            continue

        print(f"\n  Bloque {bloque} - {formatear_grupo(grupo_norm)}")

        for i in idxs:
            asignatura = str(df_total.loc[i, 'Asignatura']).strip()
            grupo = str(df_total.loc[i, 'Grupo']).strip()

            obligatorios = []
            afines = []

            for p in profesores_disponibles:
                valor = afinidades.get((p, asignatura, grupo_norm), 0)

                if valor == 2:
                    obligatorios.append(p)
                elif valor == 1:
                    afines.append(p)

            posibles = obligatorios if obligatorios else afines

            print(f"    {asignatura} - {grupo}: {posibles}")

        # Aviso importante: si dos materias del mismo desdoble solo pueden tener
        # el mismo profesor, es imposible, porque ocurren a la vez.
        for pos_a in range(len(idxs)):
            for pos_b in range(pos_a + 1, len(idxs)):
                i1 = idxs[pos_a]
                i2 = idxs[pos_b]

                asig1 = str(df_total.loc[i1, 'Asignatura']).strip()
                asig2 = str(df_total.loc[i2, 'Asignatura']).strip()

                posibles1 = []
                posibles2 = []

                for p in profesores_disponibles:
                    val1 = afinidades.get((p, asig1, grupo_norm), 0)
                    val2 = afinidades.get((p, asig2, grupo_norm), 0)

                    if val1 in [1, 2]:
                        posibles1.append(p)

                    if val2 in [1, 2]:
                        posibles2.append(p)

                if len(posibles1) == 1 and len(posibles2) == 1 and posibles1[0] == posibles2[0]:
                    print(
                        f"    [!] Posible bloqueo: {asig1} y {asig2} "
                        f"solo pueden ser impartidas por {posibles1[0]}, "
                        f"pero están en el mismo desdoble y ocurren a la vez."
                    )

    # 3. Horas seguidas
    print("\n3) Asignaturas con horas seguidas:")

    for _, fila in df_total.iterrows():
        hs = int(fila.get('HorasSeguidas', 0))

        if hs > 0:
            num_parejas = hs // 2

            print(
                f"  {fila['Asignatura']} - {fila['Grupo']}: "
                f"requiere {hs} horas seguidas "
                f"({num_parejas} pareja(s) de dos horas)."
            )

    print("\n--- FIN DEL DIAGNÓSTICO ---")


# ============================================================
# 8. FUNCIONES DE EXPORTACIÓN
# ============================================================


def profesor_asignado_a_fila(solver, x, profesores_posibles_por_fila, i):
    return next(
        (
            p for p in profesores_posibles_por_fila[i]
            if solver.Value(x[(i, p)]) == 1
        ),
        "SIN_PROFESOR"
    )



def aplicar_formato_horario(worksheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    center_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrapText=True
    )

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    fill_descanso = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    worksheet.column_dimensions['A'].width = 24

    for col in range(1, 7):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_thin

    for row in range(1, 10):
        cell = worksheet.cell(row=row, column=1)
        cell.font = Font(bold=True)
        cell.alignment = center_alignment
        cell.border = border_thin

    for col_idx in range(2, 7):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = 34

        for row_idx in range(2, 10):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = center_alignment
            cell.border = border_thin

            val = str(cell.value)

            if "RECREO" in val or "COMER" in val:
                cell.fill = fill_descanso
                cell.font = Font(bold=True)

    for row_idx in range(2, worksheet.max_row + 1):

        max_lineas = 1

        for col_idx in range(2, 7):
            valor = worksheet.cell(row=row_idx, column=col_idx).value

            if valor:
                lineas = str(valor).count("\n") + 1
                max_lineas = max(max_lineas, lineas)

        worksheet.row_dimensions[row_idx].height = max(20, 15 * max_lineas)

def aplicar_color_ocupadas(worksheet):
    fill_ocupada = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    font_normal = Font(bold=False)

    for row_idx in range(2, 10):
        for col_idx in range(2, 7):
            cell = worksheet.cell(row=row_idx, column=col_idx)

            if cell.value not in [None, ""]:
                val = str(cell.value)

                if "RECREO" not in val and "COMER" not in val:
                    cell.fill = fill_ocupada
                    cell.font = font_normal

def crear_pestana_tutores(writer, solver, x, profesores_posibles_por_fila, df_total):
    ws = writer.book.create_sheet("Tutores")

    ws.append(["Grupo", "Tutor/a"])

    grupos_individuales = sorted({
        g_ind
        for g in df_total['Grupo'].unique()
        for g_ind in expandir_grupo(g)
    })

    for g in grupos_individuales:
        tutor = ""

        for i, fila in df_total.iterrows():
            if pertenece_grupo(fila["Grupo"], g) and es_tutoria(fila["Asignatura"]):
                tutor = profesor_asignado_a_fila(
                    solver,
                    x,
                    profesores_posibles_por_fila,
                    i
                )
                break

        ws.append([g, tutor])

    aplicar_formato_tabla_simple(ws)


def crear_pestana_profesores_grupos(writer, solver, x, profesores_posibles_por_fila, df_total):
    ws = writer.book.create_sheet("Profesores-Grupos")

    profesores = sorted({
        profesor_asignado_a_fila(solver, x, profesores_posibles_por_fila, i)
        for i in df_total.index
        if profesor_asignado_a_fila(solver, x, profesores_posibles_por_fila, i) != "SIN_PROFESOR"
    })

    grupos = sorted({
        g_ind
        for g in df_total['Grupo'].unique()
        for g_ind in expandir_grupo(g)
    })

    ws.append(["Profesor"] + grupos)

    for profesor in profesores:
        fila_excel = [profesor]

        for grupo in grupos:
            imparte = False

            for i, fila in df_total.iterrows():
                prof_i = profesor_asignado_a_fila(
                    solver,
                    x,
                    profesores_posibles_por_fila,
                    i
                )

                if prof_i == profesor and pertenece_grupo(fila["Grupo"], grupo):
                    imparte = True
                    break

            fila_excel.append("X" if imparte else "")

        ws.append(fila_excel)

    aplicar_formato_tabla_simple(ws)

def aplicar_formato_tabla_resumen(worksheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrapText=True
    )

    border_thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border_thin

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = center_alignment
            cell.border = border_thin

    for col_idx in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = 28

    for row_idx in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 28


def crear_pestana_tutores(writer, solver, x, profesores_posibles_por_fila, df_total):
    datos = []

    grupos_individuales = sorted({
        g_ind
        for g in df_total["Grupo"].unique()
        for g_ind in expandir_grupo(g)
    })

    for grupo in grupos_individuales:
        tutor = ""

        for i, fila in df_total.iterrows():
            if pertenece_grupo(fila["Grupo"], grupo) and es_tutoria(fila["Asignatura"]):
                tutor = profesor_asignado_a_fila(
                    solver,
                    x,
                    profesores_posibles_por_fila,
                    i
                )
                break

        datos.append([grupo, tutor])

    df_tutores = pd.DataFrame(datos, columns=["Grupo", "Tutor/a"])
    df_tutores.to_excel(writer, sheet_name="Tutores", index=False)

    worksheet = writer.sheets["Tutores"]
    aplicar_formato_tabla_resumen(worksheet)


def crear_pestana_profesores_grupos(writer, solver, x, profesores_posibles_por_fila, df_total):
    profesores = sorted({
        profesor_asignado_a_fila(solver, x, profesores_posibles_por_fila, i)
        for i in df_total.index
        if profesor_asignado_a_fila(solver, x, profesores_posibles_por_fila, i) != "SIN_PROFESOR"
    })

    grupos = sorted({
        g_ind
        for g in df_total["Grupo"].unique()
        for g_ind in expandir_grupo(g)
    })

    datos = []

    for profesor in profesores:
        fila_excel = [profesor]

        for grupo in grupos:
            imparte = False

            for i, fila in df_total.iterrows():
                profesor_i = profesor_asignado_a_fila(
                    solver,
                    x,
                    profesores_posibles_por_fila,
                    i
                )

                if profesor_i == profesor and pertenece_grupo(fila["Grupo"], grupo):
                    imparte = True
                    break

            fila_excel.append("X" if imparte else "")

        datos.append(fila_excel)

    df_prof_grupos = pd.DataFrame(
        datos,
        columns=["Profesor"] + grupos
    )

    df_prof_grupos.to_excel(writer, sheet_name="Profesores-Grupos", index=False)

    worksheet = writer.sheets["Profesores-Grupos"]
    aplicar_formato_tabla_resumen(worksheet)

def aplicar_formato_tabla_simple(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )
    center = Alignment(horizontal="center", vertical="center", wrapText=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = border

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 28


def exportar_final(solver, v, va, x, profesores_posibles_por_fila, df_total):
    def limpiar(t):
        return str(t).replace("/", "-").replace("\\", "-")[:30]

    grupos_individuales = sorted({
        g_ind
        for g in df_total['Grupo'].unique()
        for g_ind in expandir_grupo(g)
    })

    aulas_unicas = sorted(list(set([
        a
        for lista in df_total['ListaAulas']
        for a in lista
    ])))

    h_bach = [
        "08:05-09:00",
        "09:00-09:55",
        "09:55-10:50",
        "RECREO",
        "11:20-12:15",
        "12:15-13:10",
        "13:10-14:05",
        "14:05-15:00"
    ]

    h_eso = [
        "09:00-09:55",
        "09:55-10:50",
        "RECREO (10:50-11:20)",
        "11:20-12:15",
        "12:15-13:10",
        "COMER (13:10-14:30)",
        "14:30-15:30",
        "15:30-16:30"
    ]

    # Horario común para profesores y aulas usando los slots globales del modelo
    h_global = [
        "08:05-09:00",
        "09:00-09:55",
        "09:55-10:50",
        "11:20-12:15",
        "12:15-13:10",
        "13:10-14:05",
        "14:05/14:30-15:00/15:30",
        "15:30-16:30"
    ]

    dias_nombre = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]

    with pd.ExcelWriter(ARCHIVO_SALIDA, engine='openpyxl') as writer:

        # =====================================================
        # 1. PESTAÑAS POR GRUPO
        # =====================================================
        for g_target in grupos_individuales:
            es_bach = "BACH" in normalizar_grupo(g_target)

            matriz = [["" for _ in range(5)] for _ in range(8)]

            for d in range(5):
                if es_bach:
                    matriz[3][d] = "RECREO"
                    mapping = {
                        0: 0,
                        1: 1,
                        2: 2,
                        3: 4,
                        4: 5,
                        5: 6,
                        6: 7
                    }
                else:
                    matriz[2][d] = "RECREO"
                    matriz[5][d] = "COMER"
                    mapping = {
                        1: 0,
                        2: 1,
                        3: 3,
                        4: 4,
                        6: 6,
                        7: 7
                    }

                for s_global, fila_excel in mapping.items():
                    if not slot_permitido(g_target, d, s_global):
                        matriz[fila_excel][d] = "LIBRE"
                        continue

                    items = []

                    for i, fila in df_total.iterrows():
                        if pertenece_grupo(fila['Grupo'], g_target) and solver.Value(v[(i, d, s_global)]) == 1:
                            aula = next(
                                (
                                    a for a in fila['ListaAulas']
                                    if solver.Value(va[(i, a, d, s_global)]) == 1
                                ),
                                ""
                            )

                            prof = profesor_asignado_a_fila(
                                solver,
                                x,
                                profesores_posibles_por_fila,
                                i
                            )

                            items.append(
                                f"{fila['Asignatura']}\n"
                                f"({prof})\n"
                                f"[{aula}]"
                            )

                    if items:
                        matriz[fila_excel][d] = " / ".join(items)
                    else:
                        matriz[fila_excel][d] = "SIN ASIGNAR"

            df_p = pd.DataFrame(
                matriz,
                columns=dias_nombre,
                index=h_bach if es_bach else h_eso
            )

            nombre_hoja = limpiar(g_target)
            df_p.to_excel(writer, sheet_name=nombre_hoja)

            worksheet = writer.sheets[nombre_hoja]
            aplicar_formato_horario(worksheet)

            # Lista lateral: Asignatura - Profesor - Aulas
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4F81BD",
                end_color="4F81BD",
                fill_type="solid"
            )
            center_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrapText=True
            )
            border_thin = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            col_inicio = 8

            worksheet.cell(row=1, column=col_inicio).value = "Asignatura"
            worksheet.cell(row=1, column=col_inicio + 1).value = "Profesor"
            worksheet.cell(row=1, column=col_inicio + 2).value = "Aulas posibles"

            for c in range(col_inicio, col_inicio + 3):
                cell = worksheet.cell(row=1, column=c)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border_thin

            filas_lista = []

            for i, fila in df_total.iterrows():
                if pertenece_grupo(fila['Grupo'], g_target):
                    profesor_asignado = profesor_asignado_a_fila(
                        solver,
                        x,
                        profesores_posibles_por_fila,
                        i
                    )

                    profesor_actual = str(fila['ProfesorActual']).strip()

                    # Evitamos mostrar "(antes SIN_PROFESOR)"
                    if profesor_actual == "SIN_PROFESOR" or profesor_actual == "":
                        texto_profesor = profesor_asignado
                    elif profesor_actual != profesor_asignado:
                        texto_profesor = f"{profesor_asignado}  (cambio: antes {profesor_actual})"
                    else:
                        texto_profesor = profesor_asignado

                    filas_lista.append((
                        fila['Asignatura'],
                        texto_profesor,
                        ", ".join(fila['ListaAulas'])
                    ))

            filas_lista = list(dict.fromkeys(filas_lista))

            for r, (asig, prof, aulas) in enumerate(filas_lista, start=2):
                worksheet.cell(row=r, column=col_inicio).value = asig
                worksheet.cell(row=r, column=col_inicio + 1).value = prof
                worksheet.cell(row=r, column=col_inicio + 2).value = aulas

                for c in range(col_inicio, col_inicio + 3):
                    cell = worksheet.cell(row=r, column=c)
                    cell.alignment = center_alignment
                    cell.border = border_thin

            worksheet.column_dimensions[get_column_letter(col_inicio)].width = 28
            worksheet.column_dimensions[get_column_letter(col_inicio + 1)].width = 35
            worksheet.column_dimensions[get_column_letter(col_inicio + 2)].width = 28

        # =====================================================
        # 2. PESTAÑAS POR PROFESOR
        # =====================================================
        profesores_finales = sorted({
            profesor_asignado_a_fila(solver, x, profesores_posibles_por_fila, i)
            for i in df_total.index
        })

        for profesor in profesores_finales:
            if profesor == "SIN_PROFESOR":
                continue

            matriz = [["" for _ in range(5)] for _ in range(8)]

            for d in range(5):
                for s in range(8):
                    items = []

                    for i, fila in df_total.iterrows():
                        prof_i = profesor_asignado_a_fila(
                            solver,
                            x,
                            profesores_posibles_por_fila,
                            i
                        )

                        if prof_i == profesor and solver.Value(v[(i, d, s)]) == 1:
                            aula = next(
                                (
                                    a for a in fila['ListaAulas']
                                    if solver.Value(va[(i, a, d, s)]) == 1
                                ),
                                ""
                            )

                            items.append(
                                f"{fila['Asignatura']}\n"
                                f"{fila['Grupo']}\n"
                                f"[{aula}]"
                            )

                    matriz[s][d] = " / ".join(items)

            df_prof = pd.DataFrame(
                matriz,
                columns=dias_nombre,
                index=h_global
            )

            nombre_hoja = limpiar(profesor)

            contador = 1
            nombre_original = nombre_hoja

            while nombre_hoja in writer.book.sheetnames:
                nombre_hoja = limpiar(f"{nombre_original}_{contador}")
                contador += 1

            df_prof.to_excel(writer, sheet_name=nombre_hoja)

            worksheet = writer.sheets[nombre_hoja]
            aplicar_formato_horario(worksheet)
            aplicar_color_ocupadas(worksheet)

        # =====================================================
        # 3. PESTAÑAS POR AULA
        # =====================================================
        for aula_obj in aulas_unicas:
            matriz = [["" for _ in range(5)] for _ in range(8)]

            for d in range(5):
                for s in range(8):
                    items = []

                    for i, fila in df_total.iterrows():
                        if aula_obj not in fila['ListaAulas']:
                            continue

                        if solver.Value(v[(i, d, s)]) == 1 and solver.Value(va[(i, aula_obj, d, s)]) == 1:
                            prof = profesor_asignado_a_fila(
                                solver,
                                x,
                                profesores_posibles_por_fila,
                                i
                            )

                            items.append(
                                f"{fila['Asignatura']}\n"
                                f"{fila['Grupo']}\n"
                                f"({prof})"
                            )

                    matriz[s][d] = " / ".join(items)

            df_aula = pd.DataFrame(
                matriz,
                columns=dias_nombre,
                index=h_global
            )

            nombre_hoja = limpiar("AULA_" + aula_obj)
            df_aula.to_excel(writer, sheet_name=nombre_hoja)

            worksheet = writer.sheets[nombre_hoja]
            aplicar_formato_horario(worksheet)
            aplicar_color_ocupadas(worksheet)

        # =====================================================
        # 4. PESTAÑA RESUMEN DE TUTORES
        # =====================================================
        crear_pestana_tutores(
            writer,
            solver,
            x,
            profesores_posibles_por_fila,
            df_total
        )

        # =====================================================
        # 5. PESTAÑA PROFESORES - GRUPOS
        # =====================================================
        crear_pestana_profesores_grupos(
            writer,
            solver,
            x,
            profesores_posibles_por_fila,
            df_total
        )

    print(f"Horario legible generado: {ARCHIVO_SALIDA}")



# ============================================================
# 9. FUNCIÓN PRINCIPAL: GENERACIÓN DEL HORARIO
# ============================================================

def generar_horarios():
    print("--- MOTOR V14.1: CORRECCIÓN DE RESTRICCIONES Y REJILLAS ---")

    try:
        # =====================================================
        # 1. CARGA DE ARCHIVOS Y DATOS DE ENTRADA
        # =====================================================

        xl_datos = pd.ExcelFile(ARCHIVO_DATOS)

        # Datos principales de asignaturas
        df_total = pd.read_excel(xl_datos, sheet_name="Asignaturas")

        # Matriz de afinidades profesor-asignatura-grupo
        afinidades, profesores_disponibles = cargar_afinidades(xl_datos)

        # Máximo de horas lectivas de cada profesor
        max_horas_profesor = cargar_max_horas_profesores(xl_datos)

        #Cargar los datos de los equipos
        equipos = cargar_equipos(xl_datos)

        # =====================================================
        # 2. LIMPIEZA Y NORMALIZACIÓN DE DATOS
        # =====================================================

        df_total.columns = df_total.columns.str.strip()
        df_total = df_total.fillna("")

        for col in ['Asignatura', 'Grupo', 'BloqueDesdoble', 'OpcionesAula']:
            df_total[col] = df_total[col].astype(str).str.strip()

        # Comprobación de que no haya más de un profesor obligatorio por asignatura-grupo
        comprobar_unico_profesor_obligatorio(
            df_total,
            afinidades,
            profesores_disponibles
        )

        # Número total de horas semanales de cada asignatura
        df_total['Horas'] = pd.to_numeric(
            df_total['Horas'],
            errors='coerce'
        ).fillna(0).astype(int)

        # Horas que deben impartirse de forma consecutiva.
        # 0 o vacío = sin restricción
        # 2 = una pareja de dos horas seguidas
        # 4 = dos parejas de dos horas seguidas
        if 'HorasSeguidas' not in df_total.columns:
            df_total['HorasSeguidas'] = 0

        df_total['HorasSeguidas'] = pd.to_numeric(
            df_total['HorasSeguidas'],
            errors='coerce'
        ).fillna(0).astype(int)

        # Número máximo de sesiones que una asignatura puede tener en un mismo día.
        # 0 o vacío = sin restricción específica
        # 1 = como máximo una sesión al día
        # 2 = como máximo dos sesiones al día
        if 'MaxSesionesDia' not in df_total.columns:
            df_total['MaxSesionesDia'] = 0

        df_total['MaxSesionesDia'] = pd.to_numeric(
            df_total['MaxSesionesDia'],
            errors='coerce'
        ).fillna(0).astype(int)

        # =====================================================
        # 3. PROCESAMIENTO DE AULAS POSIBLES
        # =====================================================

        def obtener_lista_aulas(fila):
            val = str(fila['OpcionesAula']).strip()

            # Si no se indica aula, se crea un aula propia del grupo.
            # Esto evita que todas las asignaturas sin aula explícita
            # compitan por una misma aula genérica.
            if val == "" or val.lower() == "nan":
                grupo_aula = normalizar_grupo(fila['Grupo'])
                return [f"Aula_{grupo_aula}"]

            return [a.strip() for a in val.split(',') if a.strip()]

        df_total['ListaAulas'] = df_total.apply(obtener_lista_aulas, axis=1)

        # =====================================================
        # 4. CARGA DE LA ASIGNACIÓN PREVIA DE PROFESORES
        # =====================================================

        xl_asig = pd.ExcelFile(ARCHIVO_ASIGNACION)
        mapeo_profes = {}

        for hoja in [
            s for s in xl_asig.sheet_names
            if s not in ["Resumen tutorías", "Carga_profesor", "Resumen", "Profesores_Clases"]
        ]:
            try:
                df_h = pd.read_excel(ARCHIVO_ASIGNACION, sheet_name=hoja).fillna("")

                if 'Asignatura' in df_h.columns:
                    for _, f in df_h.iterrows():
                        g_val = str(f.get('Clase', f.get('Grupo', ''))).strip()
                        asig_val = str(f['Asignatura']).strip()
                        profesor_val = str(f.get('Profesor', 'SIN_PROFESOR')).strip()

                        mapeo_profes[(asig_val, g_val)] = profesor_val

            except Exception as e:
                print(f"[AVISO] No se ha podido leer la hoja {hoja}: {e}")

        # Esta columna solo sirve como referencia para intentar mantener
        # la asignación anterior. No se impone como restricción dura.
        df_total['ProfesorActual'] = df_total.apply(
            lambda r: mapeo_profes.get(
                (r['Asignatura'], r['Grupo']),
                "SIN_PROFESOR"
            ),
            axis=1
        )

        # =====================================================
        # 5. CREACIÓN DEL MODELO CP-SAT Y VARIABLES DE DECISIÓN
        # =====================================================

        model = cp_model.CpModel()

        dias = range(5)     # 0 = lunes, ..., 4 = viernes
        slots = range(8)    # franjas horarias 0 a 7

        # v[(i,d,s)] = 1 si la asignatura/fila i se imparte el día d en la franja s
        v = {
            (i, d, s): model.NewBoolVar(f'v_{i}_{d}_{s}')
            for i in df_total.index
            for d in dias
            for s in slots
        }

        # Conjunto de aulas existentes
        aulas_unicas = sorted(list(set([
            a
            for lista in df_total['ListaAulas']
            for a in lista
        ])))

        # va[(i,a,d,s)] = 1 si la asignatura/fila i usa el aula a el día d en la franja s
        va = {
            (i, a, d, s): model.NewBoolVar(f'va_{i}_{a}_{d}_{s}')
            for i in df_total.index
            for a in df_total.loc[i, 'ListaAulas']
            for d in dias
            for s in slots
        }

        # x[(i,p)] = 1 si el profesor p imparte la asignatura/fila i
        x = {}

        # Diccionario auxiliar:
        # para cada fila i guarda la lista de profesores que pueden impartirla
        profesores_posibles_por_fila = {}

        # =====================================================
        # 6. CONSTRUCCIÓN DE CANDIDATOS PROFESOR-ASIGNATURA
        # =====================================================

        for i, fila in df_total.iterrows():
            asignatura = str(fila['Asignatura']).strip()
            grupo = str(fila['Grupo']).strip()
            grupo_norm = normalizar_grupo(grupo)

            profesor_actual = str(fila['ProfesorActual']).strip()

            profesores_obligatorios = []
            profesores_afines = []

            for p in profesores_disponibles:
                valor_afinidad = afinidades.get((p, asignatura, grupo_norm), 0)

                if valor_afinidad == 2:
                    profesores_obligatorios.append(p)

                elif valor_afinidad == 1:
                    profesores_afines.append(p)

            # Regla de afinidades:
            # - Si existe un único profesor con valor 2, la asignación queda fijada a él.
            # - Si no existe ningún 2, se puede elegir entre los profesores con valor 1.
            # - Si hay más de un 2, los datos de entrada son contradictorios.
            if len(profesores_obligatorios) == 1:
                posibles = profesores_obligatorios[:]

            elif len(profesores_obligatorios) > 1:
                raise ValueError(
                    f"La asignatura {asignatura} - {grupo} tiene más de un profesor "
                    f"con afinidad 2: {profesores_obligatorios}. "
                    f"Solo puede haber un profesor obligatorio."
                )

            else:
                posibles = profesores_afines[:]

            if not posibles:
                raise ValueError(
                    f"No hay ningún profesor posible para {asignatura} - {grupo}. "
                    f"En la hoja Afinidades debe haber al menos un profesor con valor 1 o 2."
                )

            # El profesor actual se usa únicamente como preferencia.
            # No se añade si no cumple las reglas de afinidad.
            if profesor_actual != "SIN_PROFESOR":
                valor_actual = afinidades.get((profesor_actual, asignatura, grupo_norm), 0)

                if valor_actual == 0:
                    print(
                        f"[AVISO] {profesor_actual} no puede impartir "
                        f"{asignatura} - {grupo} porque tiene afinidad 0. "
                        f"El modelo buscará otro profesor."
                    )

                elif valor_actual == 1 and not profesores_obligatorios:
                    if profesor_actual not in posibles:
                        posibles.append(profesor_actual)

                elif valor_actual == 2:
                    if profesor_actual not in posibles:
                        posibles.append(profesor_actual)

            profesores_posibles_por_fila[i] = posibles

            for p in posibles:
                nombre_var = f"x_{i}_{p}".replace(" ", "_").replace("/", "_")
                x[(i, p)] = model.NewBoolVar(nombre_var)

            # Cada asignatura/fila debe tener exactamente un profesor asignado.
            model.Add(sum(x[(i, p)] for p in posibles) == 1)

        # =====================================================
        # 6B. VARIABLES AUXILIARES DE OCUPACIÓN DEL PROFESOR
        # =====================================================
        # ocup_prof[(p,d,s)] = 1 si el profesor p está ocupado el día d en la sesión s.
        # Estas variables permiten:
        # - evitar solapamientos de profesores,
        # - bloquear reuniones obligatorias,
        # - buscar reuniones preferentes,
        # - intentar que tutor y cotutor coincidan.

        ocup_prof = {}
        y_prof_clase = {}

        for p in profesores_disponibles:
            for d in dias:
                for s in slots:
                    ocup_prof[(p, d, s)] = model.NewBoolVar(
                        f"ocup_{p}_{d}_{s}".replace(" ", "_").replace("/", "_")
                    )

                    variables_ocupacion = []

                    for i in df_total.index:
                        if p not in profesores_posibles_por_fila[i]:
                            continue

                        y = model.NewBoolVar(
                            f"y_prof_{i}_{p}_{d}_{s}".replace(" ", "_").replace("/", "_")
                        )

                        # y = 1 si:
                        # - la asignatura i se imparte en d,s
                        # - y además la imparte el profesor p
                        model.Add(y <= v[(i, d, s)])
                        model.Add(y <= x[(i, p)])
                        model.Add(y >= v[(i, d, s)] + x[(i, p)] - 1)

                        y_prof_clase[(i, p, d, s)] = y
                        variables_ocupacion.append(y)

                    if variables_ocupacion:
                        # Un profesor no puede tener más de una actividad simultánea.
                        model.Add(sum(variables_ocupacion) <= 1)

                        # Como ya hemos impuesto <= 1, esta suma será 0 o 1.
                        model.Add(ocup_prof[(p, d, s)] == sum(variables_ocupacion))
                    else:
                        model.Add(ocup_prof[(p, d, s)] == 0)

        # =====================================================
        # 6C. RESTRICCIONES DE EQUIPOS DE PROFESORADO
        # =====================================================

        penalizaciones_equipos = aplicar_restricciones_equipos(
            model,
            equipos,
            ocup_prof,
            profesores_disponibles,
            dias,
            slots
        )

        # =====================================================
        # 7. RESTRICCIÓN DE CARGA LECTIVA MÁXIMA POR PROFESOR
        # =====================================================

        for p in profesores_disponibles:
            max_horas = max_horas_profesor.get(p, None)

            if max_horas is None:
                raise ValueError(
                    f"No se ha encontrado MaxHoras para el profesor '{p}' "
                    f"en la hoja Profesores."
                )

            expr_carga = []

            for i, fila in df_total.iterrows():
                if p not in profesores_posibles_por_fila[i]:
                    continue

                horas_fila = int(fila['Horas'])
                expr_carga.append(horas_fila * x[(i, p)])

            if expr_carga:
                model.Add(sum(expr_carga) <= max_horas)

        # =====================================================
        # 8. RESTRICCIONES DE TUTORÍAS
        # =====================================================

        grupos_individuales_modelo = sorted({
            g_ind
            for g in df_total['Grupo'].unique()
            for g_ind in expandir_grupo(g)
        })

        # 8.1. El tutor debe impartir alguna asignatura no Tutoría en ese grupo.
        for g_obj in grupos_individuales_modelo:
            filas_tutoria = filas_tutoria_del_grupo(df_total, g_obj)

            if len(filas_tutoria) == 0:
                print(f"[AVISO] El grupo {g_obj} no tiene fila de Tutoría.")
                continue

            if len(filas_tutoria) > 1:
                print(f"[AVISO] El grupo {g_obj} tiene más de una fila de Tutoría: {filas_tutoria}")

            for i_tut in filas_tutoria:
                filas_materias_grupo = filas_no_tutoria_del_grupo(df_total, g_obj)

                for p in profesores_posibles_por_fila[i_tut]:
                    filas_mismo_profesor = [
                        i for i in filas_materias_grupo
                        if p in profesores_posibles_por_fila[i]
                    ]

                    if filas_mismo_profesor:
                        model.Add(
                            sum(x[(i, p)] for i in filas_mismo_profesor) >= x[(i_tut, p)]
                        )
                    else:
                        model.Add(x[(i_tut, p)] == 0)

        # 8.2. Cada profesor puede tener como máximo una tutoría.
        for p in profesores_disponibles:
            tutorias_posibles_p = []

            for i, fila in df_total.iterrows():
                if es_tutoria(fila['Asignatura']) and p in profesores_posibles_por_fila[i]:
                    tutorias_posibles_p.append(x[(i, p)])

            if tutorias_posibles_p:
                model.Add(sum(tutorias_posibles_p) <= 1)

        # =====================================================
        # 9. PREFERENCIA BLANDA: TUTOR CON MAYOR CARGA EN EL GRUPO
        # =====================================================

        penalizaciones_tutoria_carga = []
        PESO_TUTORIA_CARGA_GRUPO = 50

        for g_obj in grupos_individuales_modelo:
            filas_tutoria = filas_tutoria_del_grupo(df_total, g_obj)
            filas_materias_grupo = filas_no_tutoria_del_grupo(df_total, g_obj)

            for i_tut in filas_tutoria:
                for p in profesores_posibles_por_fila[i_tut]:

                    carga_real_expr = []

                    for i_mat in filas_materias_grupo:
                        if p in profesores_posibles_por_fila[i_mat]:
                            horas_mat = int(df_total.loc[i_mat, 'Horas'])
                            carga_real_expr.append(horas_mat * x[(i_mat, p)])

                    if not carga_real_expr:
                        continue

                    carga_maxima_teorica = sum(
                        int(df_total.loc[i_mat, 'Horas'])
                        for i_mat in filas_materias_grupo
                        if p in profesores_posibles_por_fila[i_mat]
                    )

                    carga_real = model.NewIntVar(
                        0,
                        carga_maxima_teorica,
                        f"carga_real_tutoria_{normalizar_grupo(g_obj)}_{i_tut}_{p}".replace(" ", "_").replace("/", "_")
                    )

                    model.Add(carga_real == sum(carga_real_expr))

                    penalizacion_baja_carga = model.NewIntVar(
                        0,
                        carga_maxima_teorica,
                        f"penal_baja_carga_{normalizar_grupo(g_obj)}_{i_tut}_{p}".replace(" ", "_").replace("/", "_")
                    )

                    # Si p es tutor, se penaliza que su carga real en el grupo sea baja.
                    model.Add(
                        penalizacion_baja_carga == carga_maxima_teorica - carga_real
                    ).OnlyEnforceIf(x[(i_tut, p)])

                    # Si p no es tutor, esta penalización no cuenta.
                    model.Add(
                        penalizacion_baja_carga == 0
                    ).OnlyEnforceIf(x[(i_tut, p)].Not())

                    penalizaciones_tutoria_carga.append(
                        PESO_TUTORIA_CARGA_GRUPO * penalizacion_baja_carga
                    )

        # =====================================================
        # 10. RESTRICCIONES DE HORAS TOTALES, ETAPA Y AULAS
        # =====================================================

        for i, fila in df_total.iterrows():
            grupo = fila['Grupo'].upper()
            es_bach = "BACH" in grupo
            es_2_bach = "2" in grupo and es_bach
            es_1_bach = "1" in grupo and es_bach

            # Cada asignatura debe impartirse exactamente el número de horas indicado.
            model.Add(
                sum(v[(i, d, s)] for d in dias for s in slots) == int(fila['Horas'])
            )

            for d in dias:
                for s in slots:
                    # Si la asignatura se imparte en una franja, debe ocupar exactamente un aula.
                    model.Add(
                        sum(va[(i, a, d, s)] for a in fila['ListaAulas']) == v[(i, d, s)]
                    )

                    # Restricciones horarias por etapa.
                    if es_bach:
                        # Bachillerato no usa el slot 7.
                        if s == 7:
                            model.Add(v[(i, d, s)] == 0)

                        # 2º Bachillerato: jueves y viernes libre en slot 6.
                        if es_2_bach and d in [3, 4] and s == 6:
                            model.Add(v[(i, d, s)] == 0)

                        # 1º Bachillerato: viernes libre en slot 6.
                        if es_1_bach and d == 4 and s == 6:
                            model.Add(v[(i, d, s)] == 0)

                    else:
                        # ESO no usa el slot 0 ni el slot 5.
                        if s in [0, 5]:
                            model.Add(v[(i, d, s)] == 0)

        # =====================================================
        # 11. MÁXIMO DE SESIONES DIARIAS POR ASIGNATURA
        # =====================================================

        for i, fila in df_total.iterrows():
            max_sesiones_dia = int(fila.get('MaxSesionesDia', 0))

            if max_sesiones_dia > 0:
                horas_seguidas = int(fila.get('HorasSeguidas', 0))

                if horas_seguidas >= 2 and max_sesiones_dia < 2:
                    raise ValueError(
                        f"Incompatibilidad en {fila['Asignatura']} - {fila['Grupo']}: "
                        f"tiene HorasSeguidas={horas_seguidas}, pero MaxSesionesDia={max_sesiones_dia}. "
                        f"Si necesita dos horas seguidas, MaxSesionesDia debe ser al menos 2."
                    )

                for d in dias:
                    model.Add(
                        sum(v[(i, d, s)] for s in slots) <= max_sesiones_dia
                    )

        # =====================================================
        # 12. SINCRONIZACIÓN DE DESDOBLES
        # =====================================================

        bloques_dict = {}

        for i, fila in df_total.iterrows():
            bloque = str(fila['BloqueDesdoble']).strip()

            if bloque == "":
                continue

            grupo_norm = normalizar_grupo(fila['Grupo'])
            clave = (grupo_norm, bloque)

            if clave not in bloques_dict:
                bloques_dict[clave] = []

            bloques_dict[clave].append(i)

        # Las asignaturas del mismo bloque de desdoble deben impartirse simultáneamente.
        for clave, idxs_bloque in bloques_dict.items():
            if len(idxs_bloque) > 1:
                rep_idx = idxs_bloque[0]

                for d in dias:
                    for s in slots:
                        for other_idx in idxs_bloque[1:]:
                            model.Add(v[(rep_idx, d, s)] == v[(other_idx, d, s)])

        # =====================================================
        # 13. PROFESORES DISTINTOS EN UN MISMO DESDOBLE
        # =====================================================

        for clave, idxs_bloque in bloques_dict.items():
            if len(idxs_bloque) <= 1:
                continue

            for p in profesores_disponibles:
                filas_con_p = [
                    i for i in idxs_bloque
                    if p in profesores_posibles_por_fila[i]
                ]

                if len(filas_con_p) > 1:
                    model.Add(
                        sum(x[(i, p)] for i in filas_con_p) <= 1
                    )

        # =====================================================
        # 14. RESTRICCIONES DE HORAS SEGUIDAS
        # =====================================================

        for i, fila in df_total.iterrows():
            horas_seguidas = int(fila['HorasSeguidas'])
            horas_totales = int(fila['Horas'])

            if horas_seguidas <= 0:
                continue

            if horas_seguidas % 2 != 0:
                raise ValueError(
                    f"HorasSeguidas debe ser un número par en "
                    f"{fila['Asignatura']} - {fila['Grupo']}. "
                    f"Valor encontrado: {horas_seguidas}"
                )

            if horas_seguidas > horas_totales:
                raise ValueError(
                    f"No se pueden pedir {horas_seguidas} horas seguidas en "
                    f"{fila['Asignatura']} - {fila['Grupo']}, porque la asignatura "
                    f"solo tiene {horas_totales} horas totales."
                )

            num_parejas_necesarias = horas_seguidas // 2

            pares_dobles = []
            pares_por_slot = {}

            for d in dias:
                for s1, s2 in pares_consecutivos_permitidos(fila['Grupo'], d):
                    y = model.NewBoolVar(f"doble_{i}_{d}_{s1}_{s2}")

                    # Si se selecciona una pareja, ambos slots deben estar ocupados.
                    model.Add(v[(i, d, s1)] == 1).OnlyEnforceIf(y)
                    model.Add(v[(i, d, s2)] == 1).OnlyEnforceIf(y)

                    pares_dobles.append(y)

                    for s in [s1, s2]:
                        clave = (d, s)
                        if clave not in pares_por_slot:
                            pares_por_slot[clave] = []
                        pares_por_slot[clave].append(y)

            if not pares_dobles:
                raise ValueError(
                    f"No hay pares consecutivos posibles para "
                    f"{fila['Asignatura']} - {fila['Grupo']}."
                )

            # Evita seleccionar dos parejas que compartan una misma franja.
            for clave, lista_y in pares_por_slot.items():
                if len(lista_y) > 1:
                    model.Add(sum(lista_y) <= 1)

            model.Add(sum(pares_dobles) >= num_parejas_necesarias)

        # =====================================================
        # 15. NO SOLAPAMIENTO DE AULAS
        # =====================================================

        for d in dias:
            for s in slots:
                for a in aulas_unicas:
                    idxs_aula = [
                        idx for idx in df_total.index
                        if a in df_total.loc[idx, 'ListaAulas']
                    ]

                    if idxs_aula:
                        model.Add(
                            sum(va[(idx, a, d, s)] for idx in idxs_aula) <= 1
                        )

        # =====================================================
        # 16. COBERTURA COMPLETA DE HUECOS LECTIVOS POR GRUPO
        # =====================================================

        for g_obj in grupos_individuales_modelo:
            representantes_g = actividades_representantes_de_grupo(df_total, g_obj)

            for d in dias:
                for s in slots:
                    if slot_permitido(g_obj, d, s):
                        expr = [v[(i, d, s)] for i in representantes_g]

                        if expr:
                            model.Add(sum(expr) == 1)

        # =====================================================
        # 17. FUNCIÓN OBJETIVO
        # =====================================================

        penalizaciones = []

        # 17.1. Penalización por cambiar la asignación previa de profesor.
        # También se favorece ligeramente la afinidad 2 frente a la afinidad 1.
        for i, fila in df_total.iterrows():
            asignatura = str(fila['Asignatura']).strip()
            grupo_norm = normalizar_grupo(fila['Grupo'])
            profesor_actual = str(fila['ProfesorActual']).strip()

            posibles = profesores_posibles_por_fila[i]

            for p in posibles:
                valor_afinidad = afinidades.get((p, asignatura, grupo_norm), 0)

                coste = 0

                if profesor_actual != "SIN_PROFESOR" and p != profesor_actual:
                    coste += 1000

                if valor_afinidad == 1:
                    coste += 10
                elif valor_afinidad == 2:
                    coste += 0

                if coste > 0:
                    penalizaciones.append(coste * x[(i, p)])

        # 17.2. Corrección importante:
        # se incorporan también las penalizaciones blandas de tutoría.
        objetivo = []

        # Penaliza cambios respecto a la asignación previa
        objetivo.extend(penalizaciones)

        # Penaliza tutorías asignadas a profesorado con poca carga en el grupo
        objetivo.extend(penalizaciones_tutoria_carga)

        # Penaliza equipos preferentes que no consiguen una hora común
        objetivo.extend(penalizaciones_equipos)

        if objetivo:
            model.Minimize(sum(objetivo))

        # =====================================================
        # 18. DIAGNÓSTICO DE HORAS POR GRUPO
        # =====================================================

        print("\n--- DIAGNÓSTICO DE HORAS POR GRUPO ---")

        for g_obj in grupos_individuales_modelo:
            representantes_g = actividades_representantes_de_grupo(df_total, g_obj)

            horas_modelo = sum(
                int(df_total.loc[i, 'Horas'])
                for i in representantes_g
            )

            huecos_disponibles = 0

            for d in dias:
                for s in slots:
                    if slot_permitido(g_obj, d, s):
                        huecos_disponibles += 1

            print(
                f"{g_obj}: {horas_modelo} horas asignadas / "
                f"{huecos_disponibles} huecos disponibles"
            )

            if horas_modelo != huecos_disponibles:
                print(
                    f"  [AVISO] Las horas de {g_obj} no coinciden con los huecos disponibles."
                )

        # =====================================================
        # 19. RESOLUCIÓN DEL MODELO
        # =====================================================

        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 180.0

        status = solver.Solve(model)

        # =====================================================
        # 20. TRATAMIENTO DE LA SOLUCIÓN
        # =====================================================

        if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
            print("--- ÉXITO ---")

            # 20.1. Carga lectiva final de cada profesor
            print("\n--- CARGA LECTIVA FINAL POR PROFESOR ---")

            for p in profesores_disponibles:
                carga = 0

                for i, fila in df_total.iterrows():
                    if p in profesores_posibles_por_fila[i] and solver.Value(x[(i, p)]) == 1:
                        carga += int(fila['Horas'])

                max_horas = max_horas_profesor.get(p, "SIN_MAX")

                if carga > 0:
                    print(f"{p}: {carga}/{max_horas} horas")

            # 20.2. Cambios realizados respecto a la asignación previa
            print("\n--- CAMBIOS DE PROFESOR REALIZADOS ---")

            hay_cambios = False

            for i, fila in df_total.iterrows():
                profesor_actual = str(fila['ProfesorActual']).strip()

                profesor_final = next(
                    (
                        p for p in profesores_posibles_por_fila[i]
                        if solver.Value(x[(i, p)]) == 1
                    ),
                    "SIN_PROFESOR"
                )

                if profesor_actual != profesor_final:
                    hay_cambios = True
                    print(
                        f"{fila['Asignatura']} - {fila['Grupo']}: "
                        f"{profesor_actual}  →  {profesor_final}"
                    )

            if not hay_cambios:
                print("No ha sido necesario cambiar ningún profesor.")

            # 20.3. Exportación final a Excel
            exportar_final(
                solver,
                v,
                va,
                x,
                profesores_posibles_por_fila,
                df_total
            )

        else:
            print("[!] Sin solución viable.")
            diagnosticar_posibles_bloqueos(
                df_total,
                afinidades,
                profesores_disponibles
            )

    except Exception as e:
        print(f"Error crítico: {e}")
