# ============================================================
# AUTORÍA DEL CÓDIGO
# Autor: Aitor Santamaria Zuluaga
# Grado en Ingeniería en Organización Industrial
# Universidad Internacional de La Rioja (UNIR)
# JULIO DEL 2026 · Trabajo Fin de Grado (TFG)
# ============================================================

import os
import random
import re
import unicodedata
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

ARCHIVO_EXCEL_DATOS = "datos_centro.xlsx"
ARCHIVO_EXCEL_SALIDA = "asignacion_completa.xlsx"

#Este es el documento de base en el que se van a concretar todos los datos para
#asignar profesores y asignaturas. Se genera en las próximas líneas.
def crear_excel_base():
    if os.path.exists(ARCHIVO_EXCEL_DATOS):
        return

    wb = Workbook()

    ws_asig = wb.active
    ws_asig.title = "Asignaturas"
    ws_asig.append(["Asignatura", "Grupo", "Horas", "MaxSesionesDia"])

    ws_prof = wb.create_sheet("Profesores")
    ws_prof.append(["Nombre", "Apellido1", "Apellido2", "MaxHoras"])

    wb.save(ARCHIVO_EXCEL_DATOS)

    print(f"Se ha creado el archivo base {ARCHIVO_EXCEL_DATOS}")
    print("Rellena las hojas 'Asignaturas' y 'Profesores' y vuelve a ejecutar.\n")

def leer_fila_encabezados(ws):
    encabezados = []
    for cell in ws[1]:
        if cell.value is None:
            encabezados.append("")
        else:
            encabezados.append(str(cell.value).strip())
    return encabezados

def normalizar(texto):
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto

def cargar_desdobles_excel():
    desdobles = {}

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "Desdobles" not in wb.sheetnames:
        return desdobles

    ws = wb["Desdobles"]

    encabezados = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in ws[1]
    ]

    idx_bloque = encabezados.index("bloque")
    idx_asignatura = encabezados.index("asignatura")
    idx_grupo = encabezados.index("grupo")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        bloque = str(row[idx_bloque]).strip()
        asignatura = str(row[idx_asignatura]).strip()
        grupo = str(row[idx_grupo]).strip()

        if bloque not in desdobles:
            desdobles[bloque] = []

        desdobles[bloque].append({
            "Asignatura": asignatura,
            "Grupo": grupo
        })

    print(f"{len(desdobles)} bloques de desdobles cargados correctamente.\n")
    return desdobles


#Función para cargar las asignaturas en el programa. En la carpeta hay un doc que 
# se puede editar para incluir nuevas asignaturas (nombre, cursoClase, horas).
def cargar_asignaturas_excel():
    asignaturas = []

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return asignaturas

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)
    
    if "Asignaturas" not in wb.sheetnames:
        print("Error: No existe la hoja 'Asignaturas' en el archivo Excel.\n")
        return asignaturas

    ws = wb["Asignaturas"]
    encabezados = leer_fila_encabezados(ws)
    cols = [c.lower() for c in encabezados]

    try:
        idx_asignatura = cols.index("asignatura")
        idx_grupo = cols.index("grupo")
        idx_horas = cols.index("horas")
        idx_max_sesiones_dia = cols.index("maxsesionesdia")
        idx_bloque_desdoble = cols.index("bloquedesdoble") if "bloquedesdoble" in cols else None
    except ValueError:
        print("Error: La hoja 'Asignaturas' debe tener las columnas 'Asignatura', 'Grupo', 'Horas', 'MaxSesionesDia'.\n")
        return asignaturas

    for num_fila, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            asignatura = str(row[idx_asignatura]).strip()
            grupo = str(row[idx_grupo]).strip()
            horas = int(row[idx_horas])
        except (ValueError, TypeError, IndexError) as e:
            print("Error en una fila de la hoja 'Asignaturas'. Se omite esa fila.")
            print(f"Fila Excel: {num_fila}")
            print(f"Contenido: {row}")
            print(f"Detalle: {e}\n")
            continue

        try:
            max_sesiones_dia = int(row[idx_max_sesiones_dia])
        except (ValueError, TypeError, IndexError):
            max_sesiones_dia = 1

        bloque_desdoble = ""

        if idx_bloque_desdoble is not None:
            if row[idx_bloque_desdoble] is not None:
                bloque_desdoble = str(row[idx_bloque_desdoble]).strip()

        asignaturas.append({
            "Asignatura": asignatura,
            "Grupo": grupo,
            "Horas": horas,
            "MaxSesionesDia": max_sesiones_dia,
            "BloqueDesdoble": bloque_desdoble
        })

    print(f"{len(asignaturas)} asignaturas cargadas correctamente desde {ARCHIVO_EXCEL_DATOS}.\n")
    return asignaturas

#Función para cargar los profesores en el programa. En la carpeta hay un CSV que 
# se puede editar para incluir nuevos profesores.
def cargar_profesores_excel():
    profesores = []

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return profesores

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "Profesores" not in wb.sheetnames:
        print("Error: No existe la hoja 'Profesores' en el archivo Excel.\n")
        return profesores

    ws = wb["Profesores"]
    encabezados = leer_fila_encabezados(ws)
    cols = [c.lower() for c in encabezados]

    try:
        idx_nombre = cols.index("nombre")
        idx_apellido1 = cols.index("apellido1")
        idx_apellido2 = cols.index("apellido2")
        idx_max_horas = cols.index("maxhoras")
    except ValueError:
        print("Error: La hoja 'Profesores' debe tener las columnas 'Nombre', 'Apellido1', 'Apellido2', 'MaxHoras'.\n")
        return profesores

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            nombre = str(row[idx_nombre]).strip()
            apellido1 = str(row[idx_apellido1]).strip()
            apellido2 = str(row[idx_apellido2]).strip()
            max_horas = int(row[idx_max_horas])
        except (TypeError, ValueError, IndexError):
            print("Error en una fila de la hoja 'Profesores'. Se omite esa fila.")
            continue

        profesores.append({
            "Nombre": nombre,
            "Apellido1": apellido1,
            "Apellido2": apellido2,
            "MaxHoras": max_horas
        })

    print(f"{len(profesores)} profesores cargados correctamente desde {ARCHIVO_EXCEL_DATOS}.\n")
    return profesores

#Función para generar un doc de profesores y asignaturas (afinidades)
def generar_hoja_afinidades_excel(profesores, asignaturas):

    if os.path.exists(ARCHIVO_EXCEL_DATOS):
        wb = load_workbook(ARCHIVO_EXCEL_DATOS)
    else:
        wb = Workbook()

    if "Afinidades" in wb.sheetnames:
        del wb["Afinidades"]

    ws = wb.create_sheet("Afinidades")

    # Fila 1: grupos
    fila_grupos = ["", "", ""]
    fila_grupos += [a["Grupo"] for a in asignaturas]

    # Fila 2: asignaturas
    fila_asignaturas = ["Nombre", "Apellido1", "Apellido2"]
    fila_asignaturas += [a["Asignatura"] for a in asignaturas]

    ws.append(fila_grupos)
    ws.append(fila_asignaturas)

    # Negrita
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cell in ws[2]:
        cell.font = Font(bold=True)

    # Filas profesores
    for profe in profesores:
        fila = [
            profe["Nombre"],
            profe["Apellido1"],
            profe["Apellido2"]
        ] + [0] * len(asignaturas)

        ws.append(fila)

    wb.save(ARCHIVO_EXCEL_DATOS)

    print("Hoja 'Afinidades' generada con fila de cursos.\n")


#Primero se genera un doc en blanco (únicamente con profesores y asignatuar),
# luego se rellenan las afinidades manualmente en la tabla (0 sin afinidad, 1 con afinidad)
# y, finalmemnte, se cargan las afinidades en el sistema.
def cargar_afinidades_excel(profesores, asignaturas):
    afinidades = []

    if not os.path.exists(ARCHIVO_EXCEL_DATOS):
        print(f"No se encontró el archivo {ARCHIVO_EXCEL_DATOS}.\n")
        return afinidades

    wb = load_workbook(ARCHIVO_EXCEL_DATOS)

    if "Afinidades" not in wb.sheetnames:
        print("No existe la hoja 'Afinidades'. Se generará automáticamente.\n")
        generar_hoja_afinidades_excel(profesores, asignaturas)
        print("Rellena las afinidades en la hoja 'Afinidades' (0 = no puede, 1 = puede, 2 = obligatorio) y vuelve a ejecutar.\n")
        return []

    ws = wb["Afinidades"]

    # fila 1 = grupos
    fila_grupos = [c.value for c in ws[1]]

    # fila 2 = asignaturas
    fila_asignaturas = [c.value for c in ws[2]]

    columnas_asignaturas = fila_asignaturas[3:]
    nombres_asignaturas = [a["Asignatura"] for a in asignaturas]

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")


    if columnas_asignaturas != nombres_asignaturas:
        print("Error: Las columnas de asignaturas no coinciden con las asignaturas cargadas.")
        return []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        fila_afinidades = []
        for valor in row[3:3 + len(nombres_asignaturas)]:
            try:
                v = int(valor)
                if v not in [0, 1, 2]:
                    v = 0
            except (TypeError, ValueError):
                v = 0
            fila_afinidades.append(v)

        afinidades.append(fila_afinidades)

    # Validación: cada columna puede tener como mucho un único 2
    for j, nombre_asignatura in enumerate(nombres_asignaturas):
        num_doses = sum(1 for i in range(len(afinidades)) if afinidades[i][j] == 2)
        if num_doses > 1:
            print(f"Error: La asignatura '{nombre_asignatura}' tiene más de un profesor marcado con 2 en la hoja 'Afinidades'.")
            print("Cada columna puede tener como mucho un único 2.\n")
            return []

    print(f"{len(afinidades)} filas de afinidades cargadas correctamente desde la hoja 'Afinidades'.\n")
    return afinidades

def profesor_da_clase_en_grupo(asignaturas, asignacion, profesor_idx, grupo_objetivo):
    grupo_objetivo_norm = normalizar_texto_grupo(grupo_objetivo)

    for j, a in enumerate(asignaturas):
        if asignacion[j] != profesor_idx:
            continue

        if "tutoria" in normalizar(a["Asignatura"]):
            continue

        grupo_asig_norm = normalizar_texto_grupo(a["Grupo"])

        if grupo_asig_norm == grupo_objetivo_norm:
            return True

        if grupo_asig_norm.endswith("AB"):
            raiz = grupo_asig_norm[:-2]
            if grupo_objetivo_norm in [raiz + "A", raiz + "B"]:
                return True

    return False


def carga_profesor_en_grupo(asignaturas, asignacion, profesor_idx, grupo_objetivo):
    grupo_objetivo_norm = normalizar_texto_grupo(grupo_objetivo)
    carga = 0

    for j, a in enumerate(asignaturas):
        if asignacion[j] != profesor_idx:
            continue

        if "tutoria" in normalizar(a["Asignatura"]):
            continue

        grupo_asig_norm = normalizar_texto_grupo(a["Grupo"])

        pertenece = grupo_asig_norm == grupo_objetivo_norm

        if grupo_asig_norm.endswith("AB"):
            raiz = grupo_asig_norm[:-2]
            pertenece = grupo_objetivo_norm in [raiz + "A", raiz + "B"]

        if pertenece:
            carga += a["Horas"]

    return carga


def normalizar_texto_grupo(grupo):
    return str(grupo).strip().upper().replace(" ", "")

#Función para asignar automáticamente un profesor a cada asigantura, teniendo en
#cuenta el número máximo de horas de cada docente y el número de horas de la asignatura.
#Partimos de las listas: profesores (nombre, ape1, ape2), asignatura (asig, grupo, horas) y
#afinidades: matriz de dos dimensiones.
def asignar_profesores_automaticamente(profesores, asignaturas, afinidades, desdobles=None):
    if desdobles is None:
        desdobles = {}

    n_profes = len(profesores)
    n_asig = len(asignaturas)


    def es_tutoria(nombre):
        return "tutoria" in normalizar(nombre)

    # 1. VALIDACIÓN PREVIA DE BLOQUES (Crucial para evitar errores de lógica)
    bloques_validar = {}
    for a in asignaturas:
        b = a.get("BloqueDesdoble", "")
        if b:
            if b not in bloques_validar:
                bloques_validar[b] = a["Horas"]
            elif bloques_validar[b] != a["Horas"]:
                print(f"Error Crítico: El bloque {b} tiene asignaturas con horas distintas ({bloques_validar[b]} vs {a['Horas']}).")
                return None

    # 2. CÁLCULO DE CARGA REAL POR GRUPO (Evita el error de >33h)
    print("--- Verificando carga real de los grupos ---")
    grupos = set(a["Grupo"] for a in asignaturas)
    for g in grupos:
        asigs_g = [a for a in asignaturas if a["Grupo"] == g]
        # Sumamos horas de lo que no tiene bloque
        horas_sueltas = sum(a["Horas"] for a in asigs_g if not a.get("BloqueDesdoble"))
        # Sumamos solo una vez las horas de cada bloque único
        bloques_g = set(a["BloqueDesdoble"] for a in asigs_g if a.get("BloqueDesdoble"))
        horas_bloques = sum(bloques_validar[b] for b in bloques_g)
        
        total_real = horas_sueltas + horas_bloques
        if total_real > 33: # O el límite de tu centro (30, 32, 33...)
            print(f"[!] ADVERTENCIA: El grupo {g} tiene {total_real}h reales. Supera el límite semanal.")

    carga_profesor = [0] * n_profes
    tutorias_profesor = [0] * n_profes
    asignacion = [-1] * n_asig
    
    # Registro de materias por clase para tutorías
    materias_por_clase = {g: [0] * n_profes for g in grupos}

    # 3. ASIGNACIÓN PRIORIZANDO BLOQUES
    # Creamos un diccionario de bloques para iterar
    dict_bloques = {}
    for j, a in enumerate(asignaturas):
        b = a.get("BloqueDesdoble", "")
        if b:
            if b not in dict_bloques: dict_bloques[b] = []
            dict_bloques[b].append(j)

    # Ordenar bloques por dificultad (más profesores requeridos primero)
    bloques_ordenados = sorted(dict_bloques.items(), key=lambda x: -len(x[1]))

    for nombre_bloque, indices in bloques_ordenados:
        profes_en_este_bloque = set()
        for j in indices:
            # Buscar profesor que:
            # 1. Tenga afinidad (1 o 2)
            # 2. Tenga horas disponibles
            # 3. NO esté ya en este mismo bloque (física elemental)
            candidatos = []
            for i in range(n_profes):
                if i in profes_en_este_bloque: continue
                if afinidades[i][j] > 0 and (carga_profesor[i] + asignaturas[j]["Horas"] <= profesores[i]["MaxHoras"]):
                    # Prioridad al '2'
                    score = afinidades[i][j] * 100 - (carga_profesor[i] / profesores[i]["MaxHoras"])
                    candidatos.append((i, score))
            
            if not candidatos:
                print(f"Error: Imposible asignar profesor a {asignaturas[j]['Asignatura']} en bloque {nombre_bloque}")
                return None
            
            # Elegir mejor candidato (mayor afinidad, menor carga)
            mejor_profe = max(candidatos, key=lambda x: x[1])[0]
            
            # Asignar
            asignacion[j] = mejor_profe
            carga_profesor[mejor_profe] += asignaturas[j]["Horas"]
            profes_en_este_bloque.add(mejor_profe)
            materias_por_clase[asignaturas[j]["Grupo"]][mejor_profe] += 1

    # 4. ASIGNAR EL RESTO (Materias normales)
    indices_normales = [
        j for j in range(n_asig)
        if asignacion[j] == -1
        and not es_tutoria(asignaturas[j]["Asignatura"])
    ]

    indices_normales.sort(
        key=lambda j: sum(
            1 for i in range(n_profes)
            if afinidades[i][j] > 0
        )
    )

    for j in indices_normales:
        
        candidatos = []
        for i in range(n_profes):
            if afinidades[i][j] > 0 and (carga_profesor[i] + asignaturas[j]["Horas"] <= profesores[i]["MaxHoras"]):
                score = afinidades[i][j] * 100 - (carga_profesor[i] / profesores[i]["MaxHoras"])
                candidatos.append((i, score))
        
        if not candidatos:
            print(f"Error: Sin profesor disponible para {asignaturas[j]['Asignatura']} - {asignaturas[j]['Grupo']}")
            return None
        
        mejor_profe = max(candidatos, key=lambda x: x[1])[0]
        asignacion[j] = mejor_profe
        carga_profesor[mejor_profe] += asignaturas[j]["Horas"]
        materias_por_clase[asignaturas[j]["Grupo"]][mejor_profe] += 1

    # 5. ASIGNAR TUTORÍAS
    indices_tutorias = [
        j for j in range(n_asig)
        if asignacion[j] == -1
        and es_tutoria(asignaturas[j]["Asignatura"])
    ]

    for j in indices_tutorias:
        grupo = asignaturas[j]["Grupo"]
        horas_tutoria = asignaturas[j]["Horas"]

        candidatos = []

        for i in range(n_profes):
            # 1. Debe tener afinidad con la tutoría
            if afinidades[i][j] <= 0:
                continue

            # 2. No puede tener ya otra tutoría
            if tutorias_profesor[i] >= 1:
                continue

            # 3. Debe impartir alguna materia en ese grupo
            if not profesor_da_clase_en_grupo(asignaturas, asignacion, i, grupo):
                continue

            # 4. Debe tener carga disponible
            if carga_profesor[i] + horas_tutoria > profesores[i]["MaxHoras"]:
                continue

            # Preferimos quien más carga tenga en el grupo.
            # En empate, quien tenga menor carga relativa global.
            carga_grupo = carga_profesor_en_grupo(asignaturas, asignacion, i, grupo)
            carga_relativa = carga_profesor[i] / profesores[i]["MaxHoras"]

            score = (
                carga_grupo * 100
                - carga_relativa
                + random.random() * 0.01
            )

            candidatos.append((i, score))

        if not candidatos:
            print(
                f"Error: Sin profesor disponible para Tutoría - {grupo}. "
                f"Debe tener afinidad, dar clase en el grupo, no tener otra tutoría "
                f"y no superar MaxHoras."
            )
            return None

        mejor_profe = max(candidatos, key=lambda x: x[1])[0]

        asignacion[j] = mejor_profe
        carga_profesor[mejor_profe] += horas_tutoria
        tutorias_profesor[mejor_profe] += 1
        materias_por_clase[grupo][mejor_profe] += horas_tutoria

    return asignacion, carga_profesor

def imprimir_tabla_asignacion(profesores, asignaturas, asignacion):
    
    #Exporta la asignación final de profesores a asignaturas en un archivo Excel.
    #Crea una hoja por grupo/clase.
    

    if asignacion is None:
        print("[!] No hay asignación que exportar.")
        return

    # Si la función recibe también la carga_profesor, nos quedamos solo con asignacion.
    if isinstance(asignacion, tuple):
        asignacion = asignacion[0]

    wb = Workbook()
    ws_inicial = wb.active
    wb.remove(ws_inicial)

    grupos = sorted(set(a["Grupo"] for a in asignaturas))

    for grupo in grupos:
        nombre_hoja = str(grupo).replace("/", "-").replace("\\", "-")[:31]
        ws = wb.create_sheet(nombre_hoja)

        ws.append(["Clase", "Asignatura", "Horas", "Profesor"])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for j, asignatura in enumerate(asignaturas):
            if asignatura["Grupo"] != grupo:
                continue

            idx_profesor = asignacion[j]

            if idx_profesor == -1:
                profesor_txt = "SIN PROFESOR"
            else:
                p = profesores[idx_profesor]
                profesor_txt = f"{p['Nombre']} {p['Apellido1']} {p['Apellido2']}"

            ws.append([
                asignatura["Grupo"],
                asignatura["Asignatura"],
                asignatura["Horas"],
                profesor_txt
            ])

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 36

    wb.save(ARCHIVO_EXCEL_SALIDA)

    print(f"Asignación exportada correctamente en {ARCHIVO_EXCEL_SALIDA}.\n")

def imprimir_carga_horaria(profesores, carga_profesor):
    
    #Imprime por consola la carga horaria final de cada profesor.
    

    print("\n--- CARGA HORARIA FINAL ---")

    for i, profesor in enumerate(profesores):
        nombre = f"{profesor['Nombre']} {profesor['Apellido1']} {profesor['Apellido2']}"
        max_horas = profesor["MaxHoras"]
        carga = carga_profesor[i]

        print(f"{nombre}: {carga}/{max_horas} horas")


def exportar_asignacion_excel(profesores, asignaturas, asignacion, carga_profesor, archivo_salida):
    """
    Exporta la asignación final a Excel, con una hoja por grupo
    y una hoja resumen de carga horaria.
    """

    if isinstance(asignacion, tuple):
        asignacion = asignacion[0]

    wb = Workbook()
    ws_inicial = wb.active
    wb.remove(ws_inicial)

    grupos = sorted(set(a["Grupo"] for a in asignaturas))

    for grupo in grupos:
        nombre_hoja = str(grupo).replace("/", "-").replace("\\", "-")[:31]
        ws = wb.create_sheet(nombre_hoja)

        ws.append(["Clase", "Asignatura", "Horas", "Profesor"])

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for j, asignatura in enumerate(asignaturas):
            if asignatura["Grupo"] != grupo:
                continue

            idx_profesor = asignacion[j]

            if idx_profesor == -1:
                profesor_txt = "SIN PROFESOR"
            else:
                p = profesores[idx_profesor]
                profesor_txt = f"{p['Nombre']} {p['Apellido1']} {p['Apellido2']}"

            ws.append([
                asignatura["Grupo"],
                asignatura["Asignatura"],
                asignatura["Horas"],
                profesor_txt
            ])

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 34
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 38

    ws_carga = wb.create_sheet("Carga_profesor")
    ws_carga.append(["Profesor", "Carga asignada", "MaxHoras", "Diferencia"])

    for cell in ws_carga[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for i, profesor in enumerate(profesores):
        nombre = f"{profesor['Nombre']} {profesor['Apellido1']} {profesor['Apellido2']}"
        carga = carga_profesor[i]
        max_horas = profesor["MaxHoras"]

        ws_carga.append([
            nombre,
            carga,
            max_horas,
            max_horas - carga
        ])

    ws_carga.column_dimensions["A"].width = 38
    ws_carga.column_dimensions["B"].width = 18
    ws_carga.column_dimensions["C"].width = 12
    ws_carga.column_dimensions["D"].width = 14

    wb.save(archivo_salida)

    print(f"Archivo de asignación exportado correctamente en {archivo_salida}.")
